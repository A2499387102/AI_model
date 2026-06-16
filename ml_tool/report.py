import os
import pandas as pd
import numpy as np
from typing import Optional
from sklearn.metrics import roc_auc_score, roc_curve, mean_squared_error, mean_absolute_error, r2_score
from scipy.stats import pearsonr, spearmanr


def _ks_auc(y_true, y_score, weights=None):
    try:
        fpr, tpr, _ = roc_curve(y_true, y_score, sample_weight=weights)
        auc = roc_auc_score(y_true, y_score, sample_weight=weights)
        if auc < 0.5:
            auc = 1 - auc
        ks = round(float(np.abs(tpr - fpr).max()), 4)
        return ks, round(float(auc), 4)
    except Exception:
        return None, None


def _mape(y_true, y_pred):
    """平均绝对百分比误差，跳过真实值为 0 的样本"""
    y_true, y_pred = np.array(y_true), np.array(y_pred)
    mask = y_true != 0
    if mask.sum() == 0:
        return None
    return round(float(np.mean(np.abs((y_true[mask] - y_pred[mask]) / y_true[mask])) * 100), 4)


def _label_rate(labels) -> Optional[float]:
    """计算二值列中 1 的占比：1值数量 / (0值数量 + 1值数量)，过滤掉非 0/1 的值。"""
    arr = np.array(labels)
    mask = (arr == 0) | (arr == 1)
    valid = arr[mask]
    if len(valid) == 0:
        return None
    return round(float(valid.sum()) / len(valid), 4)


def _build_feature_importance_df(feature_importance) -> Optional[pd.DataFrame]:
    """
    把传入的特征重要性数据整理成标准 DataFrame，供写入 Excel sheet。
    接受以下格式：
      - None / 空 → 返回 None
      - pd.DataFrame（须含 'feature' 和 'importance' 列）
      - dict {feature: importance}
      - list of (feature, importance) 元组
    返回按重要性降序排列的 DataFrame，含以下列：
      排名 | 特征名 | 重要性得分 | 占比(%) | 累计占比(%)
    """
    if feature_importance is None:
        return None

    # 统一转成 DataFrame
    if isinstance(feature_importance, dict):
        fi = pd.DataFrame(list(feature_importance.items()), columns=["feature", "importance"])
    elif isinstance(feature_importance, (list, tuple)):
        fi = pd.DataFrame(feature_importance, columns=["feature", "importance"])
    elif isinstance(feature_importance, pd.DataFrame):
        fi = feature_importance.copy()
    else:
        return None

    if fi.empty or "feature" not in fi.columns or "importance" not in fi.columns:
        return None

    fi = fi[["feature", "importance"]].copy()
    fi["importance"] = pd.to_numeric(fi["importance"], errors="coerce").fillna(0)
    fi = fi.sort_values("importance", ascending=False).reset_index(drop=True)
    fi.index += 1

    total = fi["importance"].sum()
    fi["占比(%)"]   = (fi["importance"] / total * 100).round(2) if total > 0 else 0.0
    fi["累计占比(%)"] = fi["占比(%)"].cumsum().round(2)

    fi = fi.reset_index().rename(columns={"index": "排名", "feature": "特征名", "importance": "重要性得分"})
    return fi[["排名", "特征名", "重要性得分", "占比(%)", "累计占比(%)"]]


def _write_feature_binplot_sheet(
    writer,
    raw_df: "pd.DataFrame",
    target_col: str,
    feature_importance,
    n_top: int = 20,
    n_bins: int = 5,
) -> None:
    """
    在 ExcelWriter 中生成 "特征箱线图" sheet。

    - 取特征重要性前 n_top 个特征（降序，不足则取全部）
    - 每个特征占一行，一行三张图对应 train / test / oot
    - 每张图：x 轴 = 等频 5 分箱区间，y 轴 = 箱内 target_col 均值，柱状图
    - 分箱切点统一用训练集计算，三集合共用
    - 每张图约 300×220 px，x 轴标签旋转 45 度
    - 图表标题："{feature} - {dataset名}"
    """
    import io
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    import matplotlib.font_manager as _fm
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.utils import get_column_letter

    # 中文字体：优先系统字体，Linux 上常见 Noto/WenQuanYi，找不到则退到无衬线
    _CN_CANDIDATES = [
        "Microsoft YaHei", "SimHei", "SimSun",
        "Noto Sans SC", "Noto Sans CJK SC", "Noto Sans CJK JP",
        "WenQuanYi Micro Hei", "WenQuanYi Zen Hei",
        "AR PL UMing CN", "AR PL UKai CN",
        "DejaVu Sans",  # fallback（无中文，但不乱码）
    ]
    _cn_font = None
    _available = {f.name for f in _fm.fontManager.ttflist}
    for _fn in _CN_CANDIDATES:
        if _fn in _available:
            _cn_font = _fn
            break
    if _cn_font:
        plt.rcParams["font.family"] = _cn_font
    plt.rcParams["axes.unicode_minus"] = False

    DATASET_ORDER = [
        ("train", "Train"),
        ("test",  "Test"),
        ("oot",   "OOT"),
    ]

    # 解析特征重要性，取前 n_top
    fi_df = _build_feature_importance_df(feature_importance)
    if fi_df is None:
        return
    top_features = fi_df["特征名"].tolist()[:n_top]
    if not top_features:
        return

    # 创建 sheet
    workbook = writer.book
    sheet_name = "特征箱线图"
    if sheet_name in workbook.sheetnames:
        ws = workbook[sheet_name]
    else:
        ws = workbook.create_sheet(title=sheet_name)

    IMG_W_PX = 480
    IMG_H_PX = 320
    COL_W    = IMG_W_PX / 7
    ROW_H    = IMG_H_PX / 0.75

    # 训练集，用于计算分箱切点
    train_df = raw_df[raw_df["dataset"] == "train"]

    NAME_COL  = 1
    IMG_START = 2

    for row_idx, feature in enumerate(top_features):
        if feature not in raw_df.columns:
            continue

        train_vals = train_df[feature].dropna()
        if train_vals.empty:
            continue
        try:
            quantiles = np.linspace(0, 100, n_bins + 1)
            cut_edges = np.unique(np.percentile(train_vals, quantiles))
            if len(cut_edges) < 2:
                continue
            cut_edges[0]  = -np.inf
            cut_edges[-1] =  np.inf
        except Exception:
            continue

        excel_row = row_idx + 1
        ws.cell(row=excel_row, column=NAME_COL, value=feature)
        ws.row_dimensions[excel_row].height = ROW_H
        ws.column_dimensions[get_column_letter(NAME_COL)].width = 15

        # 预计算三个集合的全局 y 轴范围，保证同一特征三张图可直接对比
        all_means_global = []
        ds_agg_cache = {}
        for ds_key, ds_label in DATASET_ORDER:
            sub_all = raw_df[raw_df["dataset"] == ds_key]
            if sub_all.empty or feature not in sub_all.columns or target_col not in sub_all.columns:
                continue
            sub_all = sub_all[[feature, target_col]].copy()
            sub     = sub_all.dropna()
            if sub.empty:
                continue
            try:
                bins_series = pd.cut(sub[feature], bins=cut_edges,
                                     include_lowest=True, duplicates="drop")
            except Exception:
                continue
            total_n = len(sub)
            agg = sub.groupby(bins_series, observed=True).agg(
                mean_target=(target_col, "mean"),
                count=(target_col, "count"),
            ).reset_index()
            agg.columns = ["bin", "mean_target", "count"]
            agg["bin_str"] = agg["bin"].astype(str)
            agg["pct"] = agg["count"] / total_n

            # Missing 箱统计（特征值为空的样本）
            sub_miss = sub_all[sub_all[feature].isna()]
            miss_mean = float(sub_miss[target_col].mean()) if not sub_miss.empty else np.nan
            miss_n    = len(sub_miss)
            miss_pct  = miss_n / len(sub_all) if len(sub_all) > 0 else 0.0

            # 数据集整体均值（含 Missing 样本，用于伪Lift基准）
            ds_overall_mean = float(sub_all[target_col].mean()) if not sub_all.empty else np.nan

            ds_agg_cache[ds_key] = (agg, miss_mean, miss_n, miss_pct, ds_overall_mean)
            all_means_global.extend(agg["mean_target"].dropna().tolist())
            if not np.isnan(miss_mean):
                all_means_global.append(miss_mean)

        if not all_means_global:
            continue
        g_ymin = float(np.min(all_means_global))
        g_ymax = float(np.max(all_means_global))
        g_pad  = (g_ymax - g_ymin) * 0.35 if (g_ymax - g_ymin) > 0 else abs(g_ymax) * 0.35 + 0.05
        y_lo, y_hi = g_ymin - g_pad * 0.15, g_ymax + g_pad

        for col_idx, (ds_key, ds_label) in enumerate(DATASET_ORDER):
            if ds_key not in ds_agg_cache:
                continue
            agg, miss_mean, miss_n, miss_pct, ds_overall_mean = ds_agg_cache[ds_key]

            # 追加 Missing 柱数据
            has_miss = miss_n > 0 and not np.isnan(miss_mean)
            bin_strs = agg["bin_str"].tolist() + (["Missing"] if has_miss else [])
            means    = np.append(agg["mean_target"].values, miss_mean) if has_miss else agg["mean_target"].values.copy()
            pcts     = np.append(agg["pct"].values, miss_pct)          if has_miss else agg["pct"].values.copy()
            n_bars   = len(bin_strs)
            x        = np.arange(n_bars)

            fig, ax_bar = plt.subplots(figsize=(IMG_W_PX / 96, IMG_H_PX / 96))
            ax_line = ax_bar.twinx()

            # 柱子：高度 = 样本占比，左轴
            bar_colors = ["#4C72B0"] * len(agg) + (["#B0784C"] if has_miss else [])
            ax_bar.bar(x, pcts, color=bar_colors, edgecolor="white", alpha=0.75, zorder=2)
            ax_bar.set_ylim(0, max(pcts.max() * 2.2, 0.1))
            ax_bar.set_ylabel("Pct", fontsize=6, color="#4C72B0")
            ax_bar.tick_params(axis="y", labelsize=5, colors="#4C72B0")

            # 折线：高度 = 均值，右轴，使用全局统一范围
            valid_x = np.arange(len(agg))
            valid_m = agg["mean_target"].values
            valid_mask_plot = ~np.isnan(valid_m)
            if valid_mask_plot.sum() > 1:
                ax_line.plot(valid_x[valid_mask_plot], valid_m[valid_mask_plot],
                             color="#E84646", linewidth=1.2, marker="o", markersize=3, zorder=3)
            if has_miss and not np.isnan(miss_mean):
                ax_line.plot(len(agg) - 1 + 1, miss_mean,
                             marker="o", markersize=3, color="#B0784C", alpha=0.5, zorder=3)
            ax_line.set_ylim(y_lo, y_hi)
            ax_line.set_ylabel("Mean", fontsize=6, color="#E84646")
            ax_line.tick_params(axis="y", labelsize=5, colors="#E84646")

            # 占比标注：柱顶上方，不依赖柱高
            pct_pad = ax_bar.get_ylim()[1] * 0.02
            for xi, pv in enumerate(pcts):
                if np.isnan(pv) or pv == 0:
                    continue
                ax_bar.text(xi, pcts[xi] + pct_pad, f"{pv:.1%}",
                            ha="center", va="bottom", fontsize=5,
                            color="#333333", zorder=4)

            # 整体均值参考线（右轴坐标）
            if not np.isnan(ds_overall_mean):
                ax_line.axhline(ds_overall_mean, color="#E84646", linewidth=0.8,
                                linestyle="--", alpha=0.6, zorder=1)

            # 均值标注：折线点正下方，边缘箱向内
            mean_pad = (y_hi - y_lo) * 0.04
            for xi, mv in enumerate(means):
                if np.isnan(mv):
                    continue
                if xi >= n_bars - 2:
                    ax_line.text(xi - 0.15, mv - mean_pad, f"{mv:.3f}",
                                 ha="right", va="top", fontsize=4.5,
                                 color="#CC2222", zorder=4)
                else:
                    ax_line.text(xi + 0.15, mv - mean_pad, f"{mv:.3f}",
                                 ha="left", va="top", fontsize=4.5,
                                 color="#CC2222", zorder=4)

            ax_bar.set_xticks(x)
            ax_bar.set_xticklabels(bin_strs, rotation=45, ha="right", fontsize=6)
            ax_bar.set_title(f"{feature} - {ds_label}", fontsize=8, pad=4)
            ax_bar.margins(x=0.05)
            fig.subplots_adjust(left=0.12, right=0.88, top=0.88, bottom=0.28)

            buf = io.BytesIO()
            fig.savefig(buf, format="png", dpi=110)
            plt.close(fig)
            buf.seek(0)

            img = XLImage(buf)
            img.width  = IMG_W_PX
            img.height = IMG_H_PX

            excel_col  = IMG_START + col_idx
            col_letter = get_column_letter(excel_col)
            ws.column_dimensions[col_letter].width = COL_W
            cell_addr = f"{col_letter}{excel_row}"
            ws.add_image(img, cell_addr)


def _make_bin_labels(mins: list, maxs: list) -> list:
    """根据每箱的实际最小值/最大值生成首尾相连的区间标签。
    相邻箱边界对齐：第 i 箱右端 = 第 i+1 箱左端（取两者边界的均值）。
    首箱左端 = 首箱最小值，末箱右端 = 末箱最大值。
    """
    n = len(mins)
    lefts  = [None] * n
    rights = [None] * n
    lefts[0]    = mins[0]
    rights[-1]  = maxs[-1]
    for i in range(n - 1):
        mid = round((maxs[i] + mins[i + 1]) / 2, 6)
        rights[i]     = mid
        lefts[i + 1]  = mid
    labels = []
    for i, (l, r) in enumerate(zip(lefts, rights)):
        lp = "(" if i > 0 else "["
        labels.append(f"{lp}{l}, {r}]")
    return labels


def _edges_to_bin_labels(edges: list) -> list:
    """把切点边界列表直接转成区间标签，首尾相连。
    edges: [e0, e1, e2, ..., en]，生成 n 个区间。
    首区间用 [ 左闭，其余用 ( 左开，所有右端用 ] 右闭。
    ±inf 原样显示为 '-inf'/'+inf'（正常情况 display_edges 传入时已替换掉）。
    """
    labels = []
    for i in range(len(edges) - 1):
        l = edges[i]
        r = edges[i + 1]
        lp = "[" if i == 0 else "("
        l_str = "-inf" if np.isinf(l) and l < 0 else ("+inf" if np.isinf(l) else str(round(l, 4)))
        r_str = "+inf" if np.isinf(r) and r > 0 else ("-inf" if np.isinf(r) else str(round(r, 4)))
        labels.append(f"{lp}{l_str}, {r_str}]")
    return labels


def _decile_table(y_true, y_score, weights=None, n_bins: int = 10) -> pd.DataFrame:
    """按预测分数等频分箱，统计每箱坏样本率、Lift、累计Lift（高→低 和 低→高）；空值样本单独成 Missing 箱"""
    total_n = len(y_true)
    total_bad = int(np.array(y_true).sum())
    overall_bad_rate = total_bad / total_n if total_n > 0 else 1e-6

    df = pd.DataFrame({"真实值": np.array(y_true), "预测分数": np.array(y_score)})
    df["权重"] = np.array(weights) if weights is not None else 1.0

    # 分离空值样本
    miss_mask = df["预测分数"].isna()
    df_valid = df[~miss_mask].copy()
    df_miss  = df[miss_mask].copy()

    try:
        df_valid["分箱"] = pd.qcut(df_valid["预测分数"], q=n_bins, duplicates="drop", labels=False) + 1
    except ValueError:
        df_valid["分箱"] = 1

    rows = []
    for bin_id, g in df_valid.groupby("分箱"):
        n = len(g)
        bad = int((g["真实值"] * g["权重"]).sum())
        bad_rate = bad / n if n > 0 else 0
        lift = round(bad_rate / overall_bad_rate, 4) if overall_bad_rate > 0 else None
        rows.append({
            "分箱": int(bin_id),
            "样本数": n,
            "样本占比": round(n / total_n, 4),
            "坏样本数": bad,
            "加权样本数": round(float(g["权重"].sum()), 2),
            "最低分": round(float(g["预测分数"].min()), 4),
            "最高分": round(float(g["预测分数"].max()), 4),
            "坏样本率": round(bad_rate, 4),
            "Lift": lift,
        })
    if not rows:
        return pd.DataFrame(columns=["分箱", "分箱区间", "样本数", "样本占比", "坏样本数",
                                     "加权样本数", "最低分", "最高分", "坏样本率", "Lift",
                                     "累计Lift(高→低)", "累计Lift(低→高)"])
    grp = pd.DataFrame(rows)
    grp.insert(1, "分箱区间", _make_bin_labels(
        grp["最低分"].tolist(), grp["最高分"].tolist()))

    # 累计Lift（高→低）：按分数从高到低累计，即从最后一箱往前累计
    cum_bad_hl  = grp["坏样本数"].iloc[::-1].cumsum().iloc[::-1].values
    cum_n_hl    = grp["样本数"].iloc[::-1].cumsum().iloc[::-1].values
    grp["累计Lift(高→低)"] = [
        round((cb / cn) / overall_bad_rate, 4) if cn > 0 else None
        for cb, cn in zip(cum_bad_hl, cum_n_hl)
    ]

    # 累计Lift（低→高）：按分数从低到高累计，即从第一箱往后累计
    cum_bad_lh = grp["坏样本数"].cumsum().values
    cum_n_lh   = grp["样本数"].cumsum().values
    grp["累计Lift(低→高)"] = [
        round((cb / cn) / overall_bad_rate, 4) if cn > 0 else None
        for cb, cn in zip(cum_bad_lh, cum_n_lh)
    ]

    # Missing 箱追加到末尾
    if not df_miss.empty:
        mn = len(df_miss)
        mbad = int((df_miss["真实值"] * df_miss["权重"]).sum())
        mbad_rate = mbad / mn if mn > 0 else 0
        mlift = round(mbad_rate / overall_bad_rate, 4) if overall_bad_rate > 0 else None
        miss_row = pd.DataFrame([{
            "分箱": "Missing",
            "分箱区间": "Missing",
            "样本数": mn,
            "样本占比": round(mn / total_n, 4),
            "坏样本数": mbad,
            "加权样本数": round(float(df_miss["权重"].sum()), 2),
            "最低分": None,
            "最高分": None,
            "坏样本率": round(mbad_rate, 4),
            "Lift": mlift,
            "累计Lift(高→低)": None,
            "累计Lift(低→高)": None,
        }])
        grp = pd.concat([grp, miss_row], ignore_index=True)

    # All 合计行（含 Missing 样本）
    all_n    = total_n
    all_bad  = int((df["真实值"] * df["权重"]).sum())
    all_br   = all_bad / all_n if all_n > 0 else 0
    all_row  = pd.DataFrame([{
        "分箱":           "All",
        "分箱区间":       "All",
        "样本数":         all_n,
        "样本占比":       1.0,
        "坏样本数":       all_bad,
        "加权样本数":     round(float(df["权重"].sum()), 2),
        "最低分":         None,
        "最高分":         None,
        "坏样本率":       round(all_br, 4),
        "Lift":           1.0,
        "累计Lift(高→低)": None,
        "累计Lift(低→高)": None,
    }])
    grp = pd.concat([grp, all_row], ignore_index=True)

    return grp


def _bucket_table(y_true, y_pred, n_bins: int = 10, bins: list = None,
                  by: str = "true", cut_edges: list = None,
                  labels: np.ndarray = None,
                  display_edges: list = None) -> pd.DataFrame:
    """按真实值或预测值分桶，统计每桶误差指标；空值样本单独成 Missing 箱。

    by:            'true' — 按真实值分桶；'pred' — 按预测值分桶
    bins:          自定义切点，传入时忽略 n_bins，自动补 ±inf
    cut_edges:     已计算好的完整 bin 边界（含 ±inf），优先级高于 bins/n_bins，用于分桶
    display_edges: 用于生成分箱区间标签的边界（±inf 已替换为训练集实际 min/max），
                   传入时所有集合共用同一套标签；不传时退回每箱实际 min/max 推导
    labels:        与 y_true 等长的二值数组（0/1），传入时在分桶结果末尾加 '1值占比' 列
    """
    total_n = len(y_true)
    df = pd.DataFrame({"真实值": np.array(y_true), "预测值": np.array(y_pred)})
    if labels is not None:
        df["_label"] = np.array(labels)
    col = "真实值" if by == "true" else "预测值"

    # 分离空值样本：分桶列或对应列任一为空均视为 missing
    miss_mask = df["真实值"].isna() | df["预测值"].isna()
    df_valid = df[~miss_mask].copy()
    df_miss  = df[miss_mask].copy()

    try:
        if cut_edges is not None:
            df_valid["分桶"] = pd.cut(df_valid[col], bins=cut_edges, labels=False, include_lowest=True)
            df_valid["分桶"] = df_valid["分桶"] + 1
        elif bins is not None:
            edges = ([-np.inf] + list(bins) + [np.inf]
                     if (bins[0] != -np.inf and bins[-1] != np.inf) else bins)
            df_valid["分桶"] = pd.cut(df_valid[col], bins=edges, labels=False, include_lowest=True)
            df_valid["分桶"] = df_valid["分桶"] + 1
        else:
            df_valid["分桶"] = pd.qcut(df_valid[col], q=n_bins, duplicates="drop", labels=False) + 1
    except ValueError:
        df_valid["分桶"] = 1

    rows = []
    for bin_id, g in df_valid.groupby("分桶"):
        n = len(g)
        mae_val  = round(float(mean_absolute_error(g["真实值"].values, g["预测值"].values)), 4)
        mape_val = _mape(g["真实值"].values, g["预测值"].values)
        row = {
            "分桶": int(bin_id),
            "样本数": n,
            "样本占比": round(n / total_n, 4),
            "真实最小值": round(float(g["真实值"].min()), 4),
            "真实最大值": round(float(g["真实值"].max()), 4),
            "真实均值": round(float(g["真实值"].mean()), 4),
            "预测均值": round(float(g["预测值"].mean()), 4),
            "预测最低值": round(float(g["预测值"].min()), 4),
            "预测最高值": round(float(g["预测值"].max()), 4),
            "MAE": mae_val,
            "MAPE(%)": mape_val,
        }
        if labels is not None:
            row["1值占比"] = _label_rate(g["_label"].values)
        rows.append(row)

    base_cols = ["分桶", "分箱区间", "样本数", "样本占比", "真实最小值", "真实最大值",
                 "真实均值", "预测均值", "预测最低值", "预测最高值",
                 "误差(预测-真实)", "MAE", "MAPE(%)"]
    if labels is not None:
        base_cols.append("1值占比")
    if not rows:
        return pd.DataFrame(columns=base_cols)
    grp = pd.DataFrame(rows)
    grp["误差(预测-真实)"] = (grp["预测均值"] - grp["真实均值"]).round(4)
    if display_edges is not None:
        bin_labels = _edges_to_bin_labels(display_edges)
        label_map = {i + 1: lbl for i, lbl in enumerate(bin_labels)}
        grp.insert(1, "分箱区间", grp["分桶"].map(lambda b: label_map.get(int(b), str(b))))
    else:
        grp.insert(1, "分箱区间", _make_bin_labels(
            grp["真实最小值"].tolist() if by == "true" else grp["预测最低值"].tolist(),
            grp["真实最大值"].tolist() if by == "true" else grp["预测最高值"].tolist()))

    # Missing 箱追加到末尾
    if not df_miss.empty:
        mn = len(df_miss)
        if df_miss["真实值"].notna().any() and df_miss["预测值"].notna().any():
            mae_m  = round(float(mean_absolute_error(df_miss["真实值"].values, df_miss["预测值"].values)), 4)
            mape_m = _mape(df_miss["真实值"].values, df_miss["预测值"].values)
            tmean  = round(float(df_miss["真实值"].mean()), 4) if df_miss["真实值"].notna().any() else None
            pmean  = round(float(df_miss["预测值"].mean()), 4) if df_miss["预测值"].notna().any() else None
        else:
            mae_m = mape_m = tmean = pmean = None
        miss_row = {
            "分桶": "Missing",
            "分箱区间": "Missing",
            "样本数": mn,
            "样本占比": round(mn / total_n, 4),
            "真实最小值": round(float(df_miss["真实值"].min()), 4) if df_miss["真实值"].notna().any() else None,
            "真实最大值": round(float(df_miss["真实值"].max()), 4) if df_miss["真实值"].notna().any() else None,
            "真实均值": tmean,
            "预测均值": pmean,
            "预测最低值": round(float(df_miss["预测值"].min()), 4) if df_miss["预测值"].notna().any() else None,
            "预测最高值": round(float(df_miss["预测值"].max()), 4) if df_miss["预测值"].notna().any() else None,
            "MAE": mae_m,
            "MAPE(%)": mape_m,
            "误差(预测-真实)": round(pmean - tmean, 4) if (pmean is not None and tmean is not None) else None,
        }
        if labels is not None:
            miss_row["1值占比"] = _label_rate(df_miss["_label"].values)
        grp = pd.concat([grp, pd.DataFrame([miss_row])], ignore_index=True)

    # All 合计行（含 Missing 样本）
    all_true  = df["真实值"].dropna()
    all_pred  = df["预测值"].dropna()
    all_n     = total_n
    if len(all_true) > 0 and len(all_pred) > 0:
        valid_both = df[df["真实值"].notna() & df["预测值"].notna()]
        all_mae  = round(float(mean_absolute_error(valid_both["真实值"].values, valid_both["预测值"].values)), 4) if not valid_both.empty else None
        all_mape = _mape(valid_both["真实值"].values, valid_both["预测值"].values) if not valid_both.empty else None
        all_tmean = round(float(all_true.mean()), 4)
        all_pmean = round(float(all_pred.mean()), 4)
    else:
        all_mae = all_mape = all_tmean = all_pmean = None
    all_row = {
        "分桶":       "All",
        "分箱区间":   "All",
        "样本数":     all_n,
        "样本占比":   1.0,
        "真实最小值": round(float(all_true.min()), 4) if len(all_true) > 0 else None,
        "真实最大值": round(float(all_true.max()), 4) if len(all_true) > 0 else None,
        "真实均值":   all_tmean,
        "预测均值":   all_pmean,
        "预测最低值": round(float(all_pred.min()), 4) if len(all_pred) > 0 else None,
        "预测最高值": round(float(all_pred.max()), 4) if len(all_pred) > 0 else None,
        "MAE":        all_mae,
        "MAPE(%)":    all_mape,
        "误差(预测-真实)": round(all_pmean - all_tmean, 4) if (all_pmean is not None and all_tmean is not None) else None,
    }
    if labels is not None:
        all_row["1值占比"] = _label_rate(df["_label"].values)
    grp = pd.concat([grp, pd.DataFrame([all_row])], ignore_index=True)

    # 统一列顺序
    grp = grp[[c for c in base_cols if c in grp.columns]]

    return grp


def _build_cross_matrix(y_true, y_pred, true_edges: list, pred_edges: list,
                        labels: np.ndarray = None,
                        display_true_edges: list = None,
                        display_pred_edges: list = None):
    """计算真实值桶 × 预测值桶交叉矩阵；真实值或预测值为空的样本单独成 Missing 行/列。

    display_true_edges / display_pred_edges：±inf 已替换为训练集实际 min/max 的边界，
    用于生成行列标签；不传则退回 _edges_to_bin_labels(true_edges/pred_edges)。
    """
    total_n = len(y_true)
    df = pd.DataFrame({"真实值": np.array(y_true), "预测值": np.array(y_pred)})
    if labels is not None:
        df["_label"] = np.array(labels)

    # 分桶（pd.cut 对 NaN 自动产生 NaN，后续用 "Missing" 特殊标记覆盖）
    df["真实桶"] = pd.cut(df["真实值"], bins=true_edges, labels=False, include_lowest=True).astype("Int64") + 1
    df["预测桶"] = pd.cut(df["预测值"], bins=pred_edges, labels=False, include_lowest=True).astype("Int64") + 1

    # 用特殊整数 -1 标记 missing（便于后续统一 groupby 逻辑）
    true_miss_mask = df["真实值"].isna()
    pred_miss_mask = df["预测值"].isna()
    MISS_ID = -1
    df.loc[true_miss_mask, "真实桶"] = MISS_ID
    df.loc[pred_miss_mask, "预测桶"] = MISS_ID

    true_bins_raw = sorted(df["真实桶"].dropna().unique())
    pred_bins_raw = sorted(df["预测桶"].dropna().unique())

    # 区分正常箱和 missing 箱，保证 missing 在末尾
    true_bins_normal = [b for b in true_bins_raw if b != MISS_ID]
    pred_bins_normal = [b for b in pred_bins_raw if b != MISS_ID]
    true_has_miss = MISS_ID in true_bins_raw
    pred_has_miss = MISS_ID in pred_bins_raw
    true_bins = true_bins_normal + ([MISS_ID] if true_has_miss else [])
    pred_bins = pred_bins_normal + ([MISS_ID] if pred_has_miss else [])

    col_label = "真实值区间\\预测值区间"

    _te = display_true_edges if display_true_edges is not None else true_edges
    _pe = display_pred_edges if display_pred_edges is not None else pred_edges
    true_labels_all = _edges_to_bin_labels(_te)
    pred_labels_all = _edges_to_bin_labels(_pe)
    true_label_map = {i + 1: lbl for i, lbl in enumerate(true_labels_all)}
    true_label_map[MISS_ID] = "Missing"
    pred_label_map = {i + 1: lbl for i, lbl in enumerate(pred_labels_all)}
    pred_label_map[MISS_ID] = "Missing"

    pred_cols = [pred_label_map.get(int(pb), f"预测桶{int(pb)}") for pb in pred_bins]

    mape_rows, count_n_rows, count_r_rows, label_rows = [], [], [], []
    for tb in true_bins:
        sub_t    = df[df["真实桶"] == tb]
        row_name = true_label_map.get(int(tb), f"真实桶{int(tb)}")
        mrow  = {col_label: row_name}
        nrow  = {col_label: row_name}
        rrow  = {col_label: row_name}
        lrow  = {col_label: row_name}
        for pb, pc in zip(pred_bins, pred_cols):
            sub = sub_t[sub_t["预测桶"] == pb]
            n   = len(sub)
            mrow[pc] = round(_mape(sub["真实值"].values, sub["预测值"].values), 2) if n else None
            nrow[pc] = n
            rrow[pc] = round(n / total_n, 4) if n else 0
            if labels is not None:
                lrow[pc] = _label_rate(sub["_label"].values)
        rt = len(sub_t)
        mrow["合计"] = round(_mape(sub_t["真实值"].values, sub_t["预测值"].values), 2) if rt else None
        nrow["合计"] = rt
        rrow["合计"] = round(rt / total_n, 4)
        if labels is not None:
            lrow["合计"] = _label_rate(sub_t["_label"].values)
        mape_rows.append(mrow)
        count_n_rows.append(nrow)
        count_r_rows.append(rrow)
        if labels is not None:
            label_rows.append(lrow)

    # 列合计行
    mrow_total = {col_label: "合计"}
    nrow_total = {col_label: "合计"}
    rrow_total = {col_label: "合计"}
    lrow_total = {col_label: "合计"}
    for pb, pc in zip(pred_bins, pred_cols):
        sub_p = df[df["预测桶"] == pb]
        np_   = len(sub_p)
        mrow_total[pc] = round(_mape(sub_p["真实值"].values, sub_p["预测值"].values), 2) if np_ else None
        nrow_total[pc] = np_
        rrow_total[pc] = round(np_ / total_n, 4)
        if labels is not None:
            lrow_total[pc] = _label_rate(sub_p["_label"].values)
    mrow_total["合计"] = round(_mape(df["真实值"].values, df["预测值"].values), 2)
    nrow_total["合计"] = total_n
    rrow_total["合计"] = 1.0
    if labels is not None:
        lrow_total["合计"] = _label_rate(df["_label"].values)
    mape_rows.append(mrow_total)
    count_n_rows.append(nrow_total)
    count_r_rows.append(rrow_total)
    if labels is not None:
        label_rows.append(lrow_total)

    label_df = pd.DataFrame(label_rows) if labels is not None else None
    return pd.DataFrame(mape_rows), pd.DataFrame(count_n_rows), pd.DataFrame(count_r_rows), label_df


def _build_gain_matrix(y_label, y_pred, score, n_bins: int = 10):
    """score × 模型预测分 交叉矩阵（各自独立等频分箱）；score 或 pred 为空时单独成 Missing 箱。

    返回 (n_df, ratio_df, bad_rate_df)：
      n_df:        格子样本数（整数）
      ratio_df:    格子样本占比（小数）
      bad_rate_df: 格子坏账率 = 逾期1样本数 / (0+1样本总和)
    行 = score 分箱，列 = 模型预测分分箱，末行/末列为合计。
    """
    total_n = len(y_label)
    df = pd.DataFrame({
        "label": np.array(y_label),
        "pred":  np.array(y_pred),
        "score": np.array(score),
    })

    MISS_ID = -1

    def _qcut_with_edges(series, bins):
        try:
            _, edges = pd.qcut(series.dropna(), q=bins, duplicates="drop", retbins=True)
            edges[0]  = -np.inf
            edges[-1] =  np.inf
            binned = pd.cut(series, bins=edges, labels=False, include_lowest=True).astype("Int64") + 1
            return binned, list(edges)
        except Exception:
            return pd.Series([1] * len(series), dtype="Int64"), [-np.inf, np.inf]

    score_binned, score_edges = _qcut_with_edges(df["score"], n_bins)
    pred_binned,  pred_edges  = _qcut_with_edges(df["pred"],  n_bins)
    df["score_bin"] = score_binned.values
    df["pred_bin"]  = pred_binned.values

    # 空值样本标记为 MISS_ID
    df.loc[df["score"].isna(), "score_bin"] = MISS_ID
    df.loc[df["pred"].isna(),  "pred_bin"]  = MISS_ID

    score_bins_raw = sorted(df["score_bin"].dropna().unique())
    pred_bins_raw  = sorted(df["pred_bin"].dropna().unique())

    score_bins_normal = [b for b in score_bins_raw if b != MISS_ID]
    pred_bins_normal  = [b for b in pred_bins_raw  if b != MISS_ID]
    score_bins = score_bins_normal + ([MISS_ID] if MISS_ID in score_bins_raw else [])
    pred_bins  = pred_bins_normal  + ([MISS_ID] if MISS_ID in pred_bins_raw  else [])

    score_lbls = _edges_to_bin_labels(score_edges)
    pred_lbls  = _edges_to_bin_labels(pred_edges)
    score_lmap = {i + 1: lbl for i, lbl in enumerate(score_lbls)}
    score_lmap[MISS_ID] = "Missing"
    pred_lmap  = {i + 1: lbl for i, lbl in enumerate(pred_lbls)}
    pred_lmap[MISS_ID]  = "Missing"
    pred_cols  = [pred_lmap.get(int(pb), str(pb)) for pb in pred_bins]
    col_label  = "score分箱\\预测分箱"

    n_rows, r_rows, br_rows = [], [], []
    for sb in score_bins:
        sub_s  = df[df["score_bin"] == sb]
        rname  = score_lmap.get(int(sb), str(sb))
        nrow   = {col_label: rname}
        rrow   = {col_label: rname}
        brrow  = {col_label: rname}
        for pb, pc in zip(pred_bins, pred_cols):
            sub = sub_s[sub_s["pred_bin"] == pb]
            n   = len(sub)
            nrow[pc]  = n
            rrow[pc]  = round(n / total_n, 4) if n else 0
            brrow[pc] = _label_rate(sub["label"].values)
        st = len(sub_s)
        nrow["合计"]  = st
        rrow["合计"]  = round(st / total_n, 4)
        brrow["合计"] = _label_rate(sub_s["label"].values)
        n_rows.append(nrow); r_rows.append(rrow); br_rows.append(brrow)

    # 列合计行
    nt = {col_label: "合计"}; rt_ = {col_label: "合计"}; brt = {col_label: "合计"}
    for pb, pc in zip(pred_bins, pred_cols):
        sub_p = df[df["pred_bin"] == pb]
        np_   = len(sub_p)
        nt[pc]  = np_
        rt_[pc] = round(np_ / total_n, 4)
        brt[pc] = _label_rate(sub_p["label"].values)
    nt["合计"]  = total_n
    rt_["合计"] = 1.0
    brt["合计"] = _label_rate(df["label"].values)
    n_rows.append(nt); r_rows.append(rt_); br_rows.append(brt)

    return pd.DataFrame(n_rows), pd.DataFrame(r_rows), pd.DataFrame(br_rows)


def _write_gain_blocks_to_sheet(ws, gain_blocks: list) -> None:
    """将增益矩阵块列表写入已创建的 worksheet，并对数据区施加条件格式。

    gain_blocks 格式：[(title_str, blk_df), ...]
    每 3 个块为一组（样本数矩阵、样本占比矩阵、坏账率矩阵），各组对应同一个数据集。
    - 样本数 / 样本占比矩阵：ColorScale（白→蓝），每组独立范围
    - 坏账率矩阵：DataBar（红色），每组独立范围
    """
    from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
    from openpyxl.utils import get_column_letter

    row_cursor = 1
    # 记录每块写入的数据区 (min_row, max_row, min_col, max_col, block_type)
    # block_type: 'n'=样本数, 'r'=样本占比, 'br'=坏账率
    block_ranges = []

    for i, (title, blk_df) in enumerate(gain_blocks):
        if i > 0:
            row_cursor += 1  # 空行
        # 标题行
        ws.cell(row=row_cursor, column=1, value=f"── {title} ──")
        row_cursor += 1
        # 列名行
        for col_idx, col_name in enumerate(blk_df.columns, 1):
            ws.cell(row=row_cursor, column=col_idx, value=col_name)
        header_row = row_cursor
        row_cursor += 1
        # 数据行（不含末列"合计"的最后一行）
        data_start_row = row_cursor
        n_data_cols = len(blk_df.columns)
        for _, data_row in blk_df.iterrows():
            for col_idx, val in enumerate(data_row, 1):
                ws.cell(row=row_cursor, column=col_idx, value=val)
            row_cursor += 1
        data_end_row = row_cursor - 1

        # 判断块类型（根据 title 关键词）
        if "坏账率" in title:
            btype = "br"
        elif "样本占比" in title:
            btype = "r"
        else:
            btype = "n"

        # 数据区：去掉第一列（行标签）和最后一行（合计行）及最后一列（合计列）
        # 取数值区域 col 2 ~ n_data_cols-1，row data_start_row ~ data_end_row-1
        if data_end_row > data_start_row and n_data_cols > 2:
            block_ranges.append((data_start_row, data_end_row - 1,
                                  2, n_data_cols - 1, btype))

    # 应用条件格式
    for (r1, r2, c1, c2, btype) in block_ranges:
        if r2 < r1 or c2 < c1:
            continue
        range_str = f"{get_column_letter(c1)}{r1}:{get_column_letter(c2)}{r2}"
        if btype == "br":
            rule = DataBarRule(
                start_type="num", start_value=0,
                end_type="max",
                color="FF0000",
            )
        else:
            rule = ColorScaleRule(
                start_type="min", start_color="FFFFFFFF",
                end_type="max",   end_color="FF4472C4",
            )
        ws.conditional_formatting.add(range_str, rule)


def _scorecard_table(score, label, n_bins: int = 10) -> pd.DataFrame:
    """标品模型分等频分箱分析，从低到高排列（第1箱分数最低）。

    列：分箱 | 分箱区间 | 样本数 | 样本占比 | 逾期数 | 逾期率 | Lift | 累计KS
    score 为空的样本单独成 Missing 箱，追加在正常箱末尾、All 行之前。
    末行追加 All 合计行（含 Missing 样本）。
    逾期率 = 逾期1样本数 / (0+1有效样本总和)
    Lift   = 箱内逾期率 / 全局逾期率
    累计KS = |累计好样本率 - 累计坏样本率|（从低到高累计，仅正常箱参与）
    """
    arr_s = np.array(score)
    arr_l = np.array(label)
    total_n = len(arr_s)

    df = pd.DataFrame({"score": arr_s, "label": arr_l})

    # 分离 Missing 样本（score 为空）
    miss_mask  = df["score"].isna()
    df_valid   = df[~miss_mask].copy()
    df_miss    = df[miss_mask].copy()

    # 全局逾期率（只用 0/1 有效样本，含 Missing 样本）
    valid_mask   = (arr_l == 0) | (arr_l == 1)
    total_bad    = int(arr_l[valid_mask].sum())
    total_good   = int((arr_l[valid_mask] == 0).sum())
    total_valid  = total_bad + total_good
    overall_bad_rate = total_bad / total_valid if total_valid > 0 else 0

    try:
        df_valid["bin"] = pd.qcut(df_valid["score"], q=n_bins, duplicates="drop", labels=False) + 1
    except Exception:
        df_valid["bin"] = 1

    rows = []
    for bin_id, g in df_valid.groupby("bin", sort=True):
        n      = len(g)
        vm     = (g["label"] == 0) | (g["label"] == 1)
        bad    = int(g.loc[vm, "label"].sum())
        good   = int((g.loc[vm, "label"] == 0).sum())
        valid  = bad + good
        br     = bad / valid if valid > 0 else None
        lift   = round(br / overall_bad_rate, 4) if (br is not None and overall_bad_rate > 0) else None
        rows.append({
            "分箱":     int(bin_id),
            "样本数":   n,
            "样本占比": round(n / total_n, 4),
            "逾期数":   bad,
            "好样本数": good,
            "逾期率":   round(br, 4) if br is not None else None,
            "Lift":     lift,
        })

    if not rows:
        return pd.DataFrame()

    grp = pd.DataFrame(rows)

    # 分箱区间
    grp.insert(1, "分箱区间", _make_bin_labels(
        df_valid.groupby("bin", sort=True)["score"].min().tolist(),
        df_valid.groupby("bin", sort=True)["score"].max().tolist(),
    ))

    # 累计KS（从低到高，仅正常箱）
    cum_bad       = grp["逾期数"].cumsum()
    cum_good      = grp["好样本数"].cumsum()
    cum_bad_rate  = cum_bad  / total_bad  if total_bad  > 0 else cum_bad  * 0
    cum_good_rate = cum_good / total_good if total_good > 0 else cum_good * 0
    grp["累计KS"] = (cum_bad_rate - cum_good_rate).abs().round(4)

    # Missing 箱（score 为空）
    if not df_miss.empty:
        mn   = len(df_miss)
        mvm  = (df_miss["label"] == 0) | (df_miss["label"] == 1)
        mbad = int(df_miss.loc[mvm, "label"].sum())
        mgood= int((df_miss.loc[mvm, "label"] == 0).sum())
        mvld = mbad + mgood
        mbr  = mbad / mvld if mvld > 0 else None
        mlift= round(mbr / overall_bad_rate, 4) if (mbr is not None and overall_bad_rate > 0) else None
        miss_row = {
            "分箱":     "Missing",
            "分箱区间": "Missing",
            "样本数":   mn,
            "样本占比": round(mn / total_n, 4),
            "逾期数":   mbad,
            "好样本数": mgood,
            "逾期率":   round(mbr, 4) if mbr is not None else None,
            "Lift":     mlift,
            "累计KS":   None,
        }
        grp = pd.concat([grp, pd.DataFrame([miss_row])], ignore_index=True)

    # All 合计行（含 Missing 样本）
    s_valid = df_valid["score"]
    all_br  = total_bad / total_valid if total_valid > 0 else None
    all_row = {
        "分箱":     "All",
        "分箱区间": f"[{s_valid.min()}, {s_valid.max()}]" if not s_valid.empty else "N/A",
        "样本数":   total_n,
        "样本占比": 1.0,
        "逾期数":   total_bad,
        "好样本数": total_good,
        "逾期率":   round(all_br, 4) if all_br is not None else None,
        "Lift":     1.0,
        "累计KS":   grp["累计KS"].dropna().max() if grp["累计KS"].notna().any() else None,
    }
    grp = pd.concat([grp, pd.DataFrame([all_row])], ignore_index=True)
    grp = grp.drop(columns=["好样本数"])
    return grp


def _monthly_stats(y_true, y_pred, months, true_edges, pred_edges,
                   labels=None, total_n_all: int = None) -> pd.DataFrame:
    """按月计算每月的 MAE、MAPE、样本数、当月占比、逾期率。

    total_n_all: 全量总样本数（不限 dataset），用于计算当月占比；
                 不传则用当前集合自身总样本数。
    """
    df = pd.DataFrame({"y": np.array(y_true), "p": np.array(y_pred), "month": np.array(months)})
    if labels is not None:
        df["_label"] = np.array(labels)

    total_base = total_n_all if total_n_all is not None else len(df)

    rows = []
    for month_val, g in df.groupby("month", sort=True):
        n = len(g)
        row = {
            "月份":     month_val,
            "样本数":   n,
            "当月占比": round(n / total_base, 4) if total_base > 0 else None,
            "MAE":      round(float(mean_absolute_error(g["y"].values, g["p"].values)), 4),
            "MAPE(%)":  _mape(g["y"].values, g["p"].values),
        }
        if labels is not None:
            row["逾期率"] = _label_rate(g["_label"].values)
        rows.append(row)

    return pd.DataFrame(rows) if rows else pd.DataFrame()


def _get_cut_edges(values, n_bins: int, bins: list) -> list:
    """根据给定数组计算分桶切点边界（含 ±inf），供多集合共享。
    bins 传入时直接转为边界，否则对 values 做等频 qcut。
    """
    if bins is not None:
        return ([-np.inf] + list(bins) + [np.inf]
                if (bins[0] != -np.inf and bins[-1] != np.inf) else list(bins))
    try:
        _, edges = pd.qcut(np.array(values), q=n_bins, duplicates="drop", retbins=True)
        edges[0]  = -np.inf
        edges[-1] = np.inf
        return list(edges)
    except ValueError:
        return [-np.inf, np.inf]


class ReportGenerator:
    """生成二分类 / 回归模型评估报告，以及特征分箱报告，输出 Excel 和 HTML

    目录结构（自动创建）：
        output_dir/
        ├── 01_feature_analysis/   特征分析报告、分箱报告
        ├── 02_feature_selection/  特征筛选报告
        ├── 03_model_tuning/       调参日志（由 ModelTrainer.fit 写入）
        ├── 04_model_report/       模型评估报告（Excel + HTML）
        └── 05_model_deploy/       模型部署文件（由 ModelTrainer.save 写入）
    """

    SUBDIRS = [
        "01_feature_analysis",
        "02_feature_selection",
        "03_model_tuning",
        "04_model_report",
        "05_model_deploy",
    ]

    def __init__(self, output_dir: str = "."):
        self.output_dir = output_dir
        for sub in self.SUBDIRS:
            os.makedirs(os.path.join(output_dir, sub), exist_ok=True)

    def _path(self, subdir: str, filename: str) -> str:
        return os.path.join(self.output_dir, subdir, filename)

    # ------------------------------------------------------------------ #
    #  二分类报告
    # ------------------------------------------------------------------ #
    def classification_report(
        self,
        datasets: dict,
        filename: str = "分类模型评估报告",
        month_col_data: Optional[dict] = None,
        feature_importance: Optional[pd.DataFrame] = None,
        gain_score_data: Optional[dict] = None,
        scorecard_data: Optional[dict] = None,
        scorecard_cols: Optional[list] = None,
        scorecard_label_cols=None,
        raw_df: Optional[pd.DataFrame] = None,
        dataset_col: str = "dataset",
    ) -> dict:
        """
        datasets:            {"数据集名": (y_true, y_score[, weight]), ...}
        month_col_data:      {"数据集名": df_with_month_col}，若为 None 则不生成按月 sheet
        feature_importance:  含 'feature' 和 'importance' 列的 DataFrame
        gain_score_data:     {"数据集名": (score, label)}，生成增益矩阵 sheet
        scorecard_cols:      df 上的标品分列名列表，与 scorecard_label_col / raw_df 配合，
                             自动拆分 train/test/oot，所有标品合并在一张 "标品分析" sheet
        scorecard_label_cols: 逾期标签列名列表（或单个列名字符串），所有标品共用
        scorecard_data:      {"标品名": {"数据集名": (score, label)}}，手动传数组方式（备用）
        """
        summary_rows = []
        sheets = {}
        decile_rows = []  # 三集合分箱合并

        for name, vals in datasets.items():
            y_true, y_score = np.array(vals[0]), np.array(vals[1])
            w = vals[2] if len(vals) > 2 else None
            ks, auc = _ks_auc(y_true, y_score, w)
            summary_rows.append({
                "数据集": name, "KS": ks, "AUC": auc,
                "样本量": len(y_true),
                "坏样本数": int(y_true.sum()),
                "坏样本率": round(float(y_true.mean()), 4),
            })
            bdf = _decile_table(y_true, y_score, w)
            bdf.insert(0, "数据集", name)
            decile_rows.append(bdf)

        if decile_rows:
            sheets["分箱分析"] = pd.concat(decile_rows, ignore_index=True)

        # 按月 KS sheet
        if month_col_data:
            ks_month_rows = []
            for name, mdf in month_col_data.items():
                if name not in datasets:
                    continue
                vals = datasets[name]
                y_true, y_score = np.array(vals[0]), np.array(vals[1])
                mdf = mdf.reset_index(drop=True)
                month_col = _detect_month_col(mdf)
                if month_col is None:
                    continue
                tmp = pd.DataFrame({"y": y_true, "s": y_score, "month": mdf[month_col].values})
                for month_val, g in tmp.groupby("month", sort=True):
                    if g["y"].nunique() < 2:
                        continue
                    ks_m, auc_m = _ks_auc(g["y"].values, g["s"].values)
                    ks_month_rows.append({
                        "数据集": name, month_col: month_val,
                        "样本数": len(g), "坏样本数": int(g["y"].sum()),
                        "坏样本率": round(float(g["y"].mean()), 4),
                        "KS": ks_m, "AUC": auc_m,
                    })
            if ks_month_rows:
                sheets["按月KS_AUC"] = pd.DataFrame(ks_month_rows)

            # 按月分箱：三集合合并为一张 sheet
            month_bin_all = []
            for name, mdf in month_col_data.items():
                if name not in datasets:
                    continue
                vals = datasets[name]
                y_true, y_score = np.array(vals[0]), np.array(vals[1])
                mdf = mdf.reset_index(drop=True)
                month_col = _detect_month_col(mdf)
                if month_col is None:
                    continue
                tmp = pd.DataFrame({"y": y_true, "s": y_score, "month": mdf[month_col].values})
                for month_val, g in tmp.groupby("month", sort=True):
                    bin_df = _decile_table(g["y"].values, g["s"].values)
                    bin_df.insert(0, month_col, month_val)
                    bin_df.insert(0, "数据集", name)
                    month_bin_all.append(bin_df)
            if month_bin_all:
                sheets["按月分箱"] = pd.concat(month_bin_all, ignore_index=True)

        # 增益矩阵（可选）—— 存为 [(title, df), ...] 列表，写 Excel 时逐块手动控制行
        gain_blocks_clf = []  # [(title_str, df), ...]
        if gain_score_data:
            for name, (score, g_label) in gain_score_data.items():
                if name not in datasets:
                    continue
                y_score_gain = np.array(datasets[name][1])
                n_df_g, ratio_df_g, br_df_g = _build_gain_matrix(
                    np.array(g_label), y_score_gain, np.array(score)
                )
                gain_blocks_clf.append((f"{name} — 样本数矩阵（行=score分箱，列=预测分分箱）", n_df_g))
                gain_blocks_clf.append((f"{name} — 样本占比矩阵", ratio_df_g))
                gain_blocks_clf.append((f"{name} — 坏账率矩阵", br_df_g))


        # 标品模型分分析（可选）
        # 标品模型分分析（可选）
        # 优先从 raw_df 按列名构建，否则用手动传入的 scorecard_data
        # scorecard_label_cols 支持单个字符串或列表
        name_map_rev = {"训练集": "train", "验证集": "test", "OOT": "oot"}
        _label_cols = (
            [scorecard_label_cols] if isinstance(scorecard_label_cols, str)
            else (scorecard_label_cols or [])
        )
        if scorecard_cols and _label_cols and raw_df is not None:
            # 结构：{(标品列名, 逾期标签列名): {"数据集名": (score, label)}}
            scorecard_data = {}
            for col in scorecard_cols:
                for lbl_col in _label_cols:
                    ds_dict = {}
                    for ds_name in datasets:
                        ds_key = name_map_rev.get(ds_name, ds_name)
                        sub = raw_df[raw_df[dataset_col] == ds_key].reset_index(drop=True)
                        if sub.empty or col not in sub.columns or lbl_col not in sub.columns:
                            continue
                        ds_dict[ds_name] = (sub[col].values, sub[lbl_col].values)
                    if ds_dict:
                        scorecard_data[(col, lbl_col)] = ds_dict

        if scorecard_data:
            all_card_rows = []
            for (card_name, lbl_col), ds_dict in scorecard_data.items():
                for ds_name, (sc_score, sc_label) in ds_dict.items():
                    sc_df = _scorecard_table(sc_score, sc_label)
                    if sc_df.empty:
                        continue
                    sc_df.insert(0, "数据集", ds_name)
                    sc_df.insert(0, "逾期标签", lbl_col)
                    sc_df.insert(0, "标品名称", card_name)
                    all_card_rows.append(sc_df)
            if all_card_rows:
                sheets["标品分析"] = pd.concat(all_card_rows, ignore_index=True)

        summary_df = pd.DataFrame(summary_rows)
        excel_path = self._path("04_model_report", f"{filename}.xlsx")
        html_path  = self._path("04_model_report", f"{filename}.html")

        with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
            summary_df.to_excel(writer, sheet_name="汇总指标", index=False)
            for sname, sdf in sheets.items():
                sdf.to_excel(writer, sheet_name=sname[:31], index=False)
            fi_df = _build_feature_importance_df(feature_importance)
            if fi_df is not None:
                fi_df.to_excel(writer, sheet_name="特征重要性", index=False)
            # 增益矩阵：逐块写入，每块标题行 + 列名行 + 数据行，不写顶层 header
            if gain_blocks_clf:
                ws = writer.book.create_sheet("增益矩阵")
                _write_gain_blocks_to_sheet(ws, gain_blocks_clf)

        html = self._build_clf_html(summary_df, sheets, filename)
        with open(html_path, "w", encoding="utf-8") as f:
            f.write(html)

        return {"summary": summary_df, "sheets": sheets,
                "excel_path": excel_path, "html_path": html_path}

    # ------------------------------------------------------------------ #
    #  回归报告
    # ------------------------------------------------------------------ #
    def regression_report(
        self,
        datasets: dict,
        filename: str = "回归模型评估报告",
        n_bins: int = 10,
        bins: list = None,
        n_bins_pred: int = 10,
        bins_pred: list = None,
        month_col_data: Optional[dict] = None,
        label_col_data: Optional[dict] = None,
        feature_importance: Optional[pd.DataFrame] = None,
        raw_df: Optional[pd.DataFrame] = None,
        target_col: Optional[str] = None,
        gain_score_data: Optional[dict] = None,
        gain_score_col: Optional[str] = None,
        gain_label_col: Optional[str] = None,
        scorecard_data: Optional[dict] = None,
        scorecard_cols: Optional[list] = None,
        scorecard_label_cols=None,
    ) -> dict:
        """
        datasets:         {"数据集名": (y_true, y_pred), ...}
                          第一个数据集视为训练集，用于计算共享切点。
        n_bins:           真实值等频分桶数（bins 未传时生效），默认 10
        bins:             真实值自定义切点（传入时忽略 n_bins）
        n_bins_pred:      预测值等频分桶数（bins_pred 未传时生效），默认 10
        bins_pred:        预测值自定义切点（传入时忽略 n_bins_pred）
        label_col_data:   {"数据集名": array_like}，各集合对应的二值列（0/1）
        gain_score_data:  {"数据集名": (score, label)}，各集合的外部 score 和逾期标签（手动传数组）
        gain_score_col:   raw_df 中的 score 列名，与 gain_label_col 配合使用，
                          自动按 dataset 列拆分，优先级高于 gain_score_data
        gain_label_col:   raw_df 中的逾期标签列名（0/1），与 gain_score_col 配合使用
        以上两种方式传入任意一种均可生成 "增益矩阵" sheet
        """
        summary_rows = []
        sheets = {}

        # 用第一个数据集（训练集）计算两套共享切点
        first_vals = list(datasets.values())[0]
        true_edges = _get_cut_edges(np.array(first_vals[0]), n_bins,      bins)
        pred_edges = _get_cut_edges(np.array(first_vals[1]), n_bins_pred, bins_pred)

        # display_edges：头尾保留 ±inf（标签显示为 -inf/+inf，明确开放边界），
        # 中间切点直接用 _get_cut_edges 计算的训练集 qcut 值，三集合共用同一套标签
        display_true_edges = list(true_edges)
        display_pred_edges = list(pred_edges)

        merged_true_rows = []
        merged_pred_rows = []
        merged_matrix_rows = []
        for name, (y_true, y_pred) in datasets.items():
            y_true, y_pred = np.array(y_true), np.array(y_pred)
            lbl = np.array(label_col_data[name]) if label_col_data and name in label_col_data else None

            rmse = round(float(np.sqrt(mean_squared_error(y_true, y_pred))), 4)
            mae  = round(float(mean_absolute_error(y_true, y_pred)), 4)
            r2   = round(float(r2_score(y_true, y_pred)), 4)
            mape = _mape(y_true, y_pred)
            pearson_r,  pearson_p  = pearsonr(y_true, y_pred)
            spearman_r, spearman_p = spearmanr(y_true, y_pred)
            summary_rows.append({
                "数据集": name, "RMSE": rmse, "MAE": mae, "R2": r2, "MAPE(%)": mape,
                "Pearson相关系数":  round(float(pearson_r),  4),
                "Pearson_p值":      round(float(pearson_p),  4),
                "Spearman相关系数": round(float(spearman_r), 4),
                "Spearman_p值":     round(float(spearman_p), 4),
                "样本量": len(y_true),
            })
            bdf_t = _bucket_table(y_true, y_pred, by="true", cut_edges=true_edges, labels=lbl,
                                  display_edges=display_true_edges)
            bdf_t.insert(0, "数据集", name)
            merged_true_rows.append(bdf_t)

            bdf_p = _bucket_table(y_true, y_pred, by="pred", cut_edges=pred_edges, labels=lbl,
                                  display_edges=display_pred_edges)
            bdf_p.insert(0, "数据集", name)
            merged_pred_rows.append(bdf_p)

            merged_matrix_rows.append((name, _build_cross_matrix(
                y_true, y_pred, true_edges, pred_edges, labels=lbl,
                display_true_edges=display_true_edges,
                display_pred_edges=display_pred_edges)))

        # 三张 sheet：真实值分桶、预测值分桶、分桶矩阵，各自合并三集合
        def _section_header(title: str, cols: list) -> pd.DataFrame:
            row = {c: "" for c in cols}
            row[cols[0]] = f"── {title} ──"
            return pd.DataFrame([row])

        def _blank_row(cols: list) -> pd.DataFrame:
            return pd.DataFrame([{c: "" for c in cols}])

        if merged_true_rows:
            sheets["真实值分桶"] = pd.concat(merged_true_rows, ignore_index=True)
        if merged_pred_rows:
            sheets["预测值分桶"] = pd.concat(merged_pred_rows, ignore_index=True)

        # 分桶矩阵 sheet：三集合合并，含 MAPE + 样本数 + 样本占比 + 1值占比（可选）
        mat_parts = []
        for ds_name, (mape_df, n_df, ratio_df, label_df) in merged_matrix_rows:
            if mat_parts:
                mat_parts.append(_blank_row(mape_df.columns.tolist()))
            mat_parts.append(_section_header(f"{ds_name} — MAPE矩阵（行=真实值桶，列=预测值桶）", mape_df.columns.tolist()))
            mat_parts.append(mape_df)
            mat_parts.append(_blank_row(n_df.columns.tolist()))
            mat_parts.append(_section_header(f"{ds_name} — 样本数矩阵", n_df.columns.tolist()))
            mat_parts.append(n_df)
            mat_parts.append(_blank_row(ratio_df.columns.tolist()))
            mat_parts.append(_section_header(f"{ds_name} — 样本占比矩阵", ratio_df.columns.tolist()))
            mat_parts.append(ratio_df)
            if label_df is not None:
                mat_parts.append(_blank_row(label_df.columns.tolist()))
                mat_parts.append(_section_header(f"{ds_name} — 1值占比矩阵", label_df.columns.tolist()))
                mat_parts.append(label_df)
        if mat_parts:
            sheets["分桶矩阵"] = pd.concat(mat_parts, ignore_index=True)

        # 增益矩阵 sheet（可选）
        # 优先从 raw_df 按列名构建，否则用手动传入的 gain_score_data
        name_map_rev = {"训练集": "train", "验证集": "test", "OOT": "oot"}
        if gain_score_col and gain_label_col and raw_df is not None:
            ds_col = "dataset"
            gain_score_data = {}
            for ds_name in datasets:
                ds_key = name_map_rev.get(ds_name, ds_name)
                sub = raw_df[raw_df[ds_col] == ds_key].reset_index(drop=True)
                if sub.empty or gain_score_col not in sub.columns or gain_label_col not in sub.columns:
                    continue
                gain_score_data[ds_name] = (sub[gain_score_col].values, sub[gain_label_col].values)

        # 增益矩阵（可选）—— 存为 [(title, df), ...] 列表，写 Excel 时逐块手动控制行
        gain_blocks_reg = []
        if gain_score_data:
            for name, (score, g_label) in gain_score_data.items():
                if name not in datasets:
                    continue
                y_pred_gain = np.array(datasets[name][1])
                n_df_g, ratio_df_g, br_df_g = _build_gain_matrix(
                    np.array(g_label), y_pred_gain, np.array(score)
                )
                gain_blocks_reg.append((f"{name} — 样本数矩阵（行=score分箱，列=预测分分箱）", n_df_g))
                gain_blocks_reg.append((f"{name} — 样本占比矩阵", ratio_df_g))
                gain_blocks_reg.append((f"{name} — 坏账率矩阵", br_df_g))

        # 标品模型分分析（可选）
        # 优先从 raw_df 按列名构建
        name_map_rev_reg = {"训练集": "train", "验证集": "test", "OOT": "oot"}
        _label_cols_reg = (
            [scorecard_label_cols] if isinstance(scorecard_label_cols, str)
            else (scorecard_label_cols or [])
        )
        if scorecard_cols and _label_cols_reg and raw_df is not None:
            scorecard_data = {}
            for col in scorecard_cols:
                for lbl_col in _label_cols_reg:
                    ds_dict = {}
                    for ds_name in datasets:
                        ds_key = name_map_rev_reg.get(ds_name, ds_name)
                        sub = raw_df[raw_df["dataset"] == ds_key].reset_index(drop=True)
                        if sub.empty or col not in sub.columns or lbl_col not in sub.columns:
                            continue
                        ds_dict[ds_name] = (sub[col].values, sub[lbl_col].values)
                    if ds_dict:
                        scorecard_data[(col, lbl_col)] = ds_dict

        if scorecard_data:
            all_card_rows = []
            for (card_name, lbl_col), ds_dict in scorecard_data.items():
                for ds_name, (sc_score, sc_label) in ds_dict.items():
                    sc_df = _scorecard_table(sc_score, sc_label)
                    if sc_df.empty:
                        continue
                    sc_df.insert(0, "数据集", ds_name)
                    sc_df.insert(0, "逾期标签", lbl_col)
                    sc_df.insert(0, "标品名称", card_name)
                    all_card_rows.append(sc_df)
            if all_card_rows:
                sheets["标品分析"] = pd.concat(all_card_rows, ignore_index=True)

        # 按月分桶 sheet + 按月评估（合并三集合）
        monthly_summary_rows = []
        # 全量总样本数（所有 dataset 合并），用于按月占比分母
        total_n_all = sum(len(np.array(v[0])) for v in datasets.values())
        if month_col_data:
            for name, mdf in month_col_data.items():
                if name not in datasets:
                    continue
                y_true_all, y_pred_all = np.array(datasets[name][0]), np.array(datasets[name][1])
                lbl_all = np.array(label_col_data[name]) if label_col_data and name in label_col_data else None
                mdf = mdf.reset_index(drop=True)
                month_col = _detect_month_col(mdf)
                if month_col is None:
                    continue
                tmp = pd.DataFrame({"y": y_true_all, "p": y_pred_all,
                                    "month": mdf[month_col].values})
                if lbl_all is not None:
                    tmp["_label"] = lbl_all

                # 按月汇总指标（每月一行）
                ms_df = _monthly_stats(
                    y_true_all, y_pred_all, mdf[month_col].values,
                    true_edges, pred_edges,
                    labels=lbl_all,
                    total_n_all=total_n_all,
                )
                if not ms_df.empty:
                    ms_df.insert(0, "数据集", name)
                    monthly_summary_rows.append(ms_df)

        if monthly_summary_rows:
            sheets["按月评估"] = pd.concat(monthly_summary_rows, ignore_index=True)

        summary_df = pd.DataFrame(summary_rows)
        excel_path = self._path("04_model_report", f"{filename}.xlsx")
        html_path  = self._path("04_model_report", f"{filename}.html")

        with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
            summary_df.to_excel(writer, sheet_name="汇总指标", index=False)
            for sname, sdf in sheets.items():
                sdf.to_excel(writer, sheet_name=sname[:31], index=False)
            fi_df = _build_feature_importance_df(feature_importance)
            if fi_df is not None:
                fi_df.to_excel(writer, sheet_name="特征重要性", index=False)
            # 真实值分桶 / 预测值分桶：MAPE(%)→ColorScale白→蓝，1值占比→DataBar红
            from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
            from openpyxl.utils import get_column_letter
            for _sname in ["真实值分桶", "预测值分桶"]:
                if _sname not in writer.book.sheetnames:
                    continue
                _ws = writer.book[_sname]
                _header = {str(_ws.cell(row=1, column=c).value): c
                           for c in range(1, _ws.max_column + 1)}
                _col_mape  = _header.get("MAPE(%)")
                _col_label = _header.get("1值占比")
                _data_rows = []
                for r in range(2, _ws.max_row + 1):
                    v = str(_ws.cell(row=r, column=1).value or "")
                    if "──" not in v and v != "None" and v != "":
                        _data_rows.append(r)
                if not _data_rows:
                    continue
                r1, r2 = _data_rows[0], _data_rows[-1]
                if _col_mape:
                    _cl = get_column_letter(_col_mape)
                    _ws.conditional_formatting.add(
                        f"{_cl}{r1}:{_cl}{r2}",
                        ColorScaleRule(start_type="min", start_color="FFFFFFFF",
                                       end_type="max", end_color="FF4472C4"))
                if _col_label:
                    _cl = get_column_letter(_col_label)
                    _ws.conditional_formatting.add(
                        f"{_cl}{r1}:{_cl}{r2}",
                        DataBarRule(start_type="num", start_value=0,
                                    end_type="max", color="FF0000"))
            # 分桶矩阵条件格式：MAPE→DataBar红，样本数/样本占比→ColorScale白→蓝
            if "分桶矩阵" in writer.book.sheetnames:
                from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
                from openpyxl.utils import get_column_letter
                ws_mat = writer.book["分桶矩阵"]
                max_col = ws_mat.max_column
                # 扫描标题行，识别各块数据区域
                i = 1
                while i <= ws_mat.max_row:
                    cell_val = str(ws_mat.cell(row=i, column=1).value or "")
                    is_mape  = "MAPE矩阵" in cell_val
                    is_n     = "样本数矩阵" in cell_val
                    is_ratio = "样本占比矩阵" in cell_val
                    if is_mape or is_n or is_ratio:
                        # 下一行是列名行，再下一行开始是数据
                        data_start = i + 2
                        data_end   = data_start
                        while data_end <= ws_mat.max_row:
                            next_val = str(ws_mat.cell(row=data_end + 1, column=1).value or "")
                            if next_val == "" or "矩阵" in next_val or "──" in next_val:
                                break
                            data_end += 1
                        if data_end >= data_start:
                            # 数据区：排除第1列（行标签）和最后一列（合计）
                            c1 = get_column_letter(2)
                            c2 = get_column_letter(max(2, max_col - 1))
                            range_str = f"{c1}{data_start}:{c2}{data_end}"
                            if is_mape:
                                rule = DataBarRule(
                                    start_type="num", start_value=0,
                                    end_type="max", color="FF0000")
                            else:
                                rule = ColorScaleRule(
                                    start_type="min", start_color="FFFFFFFF",
                                    end_type="max", end_color="FF4472C4")
                            ws_mat.conditional_formatting.add(range_str, rule)
                        i = data_end + 1
                    else:
                        i += 1
            # 增益矩阵：逐块写入，每块标题行 + 列名行 + 数据行
            if gain_blocks_reg:
                ws = writer.book.create_sheet("增益矩阵")
                _write_gain_blocks_to_sheet(ws, gain_blocks_reg)
            # 特征箱线图 sheet
            if feature_importance is not None and raw_df is not None and target_col is not None:
                _write_feature_binplot_sheet(
                    writer, raw_df, target_col, feature_importance
                )

        html = self._build_reg_html(summary_df, sheets, filename)
        with open(html_path, "w", encoding="utf-8") as f:
            f.write(html)

        return {"summary": summary_df, "sheets": sheets,
                "excel_path": excel_path, "html_path": html_path}

    # ------------------------------------------------------------------ #
    #  DataFrame 入口 — 分类
    # ------------------------------------------------------------------ #
    def classification_report_from_df(
        self,
        df: pd.DataFrame,
        target_col: str,
        pred_col: str,
        dataset_col: str = "dataset",
        month_col: Optional[str] = None,
        weight_col: Optional[str] = None,
        filename: str = "分类模型评估报告",
        feature_importance: Optional[pd.DataFrame] = None,
        gain_score_col: Optional[str] = None,
        gain_label_col: Optional[str] = None,
        scorecard_cols: Optional[list] = None,
        scorecard_label_cols=None,
        scorecard_data: Optional[dict] = None,
    ) -> dict:
        """
        直接传入完整 DataFrame，内部按 dataset_col 拆分为 train / test / oot 三个集合。

        dataset_col 的值约定：'train' = 训练集，'test' = 验证集，'oot' = OOT
        month_col:           月份列名，传入后生成按月 KS/AUC 和按月分箱 sheet
        weight_col:          样本权重列名，不传则等权
        gain_score_col:      score 列名，与 gain_label_col 配合生成增益矩阵 sheet
        gain_label_col:      逾期标签列名（0/1）
        scorecard_cols:      标品分列名列表，与 scorecard_label_col 配合生成标品分析 sheet
        scorecard_label_cols: 逾期标签列名列表（或单个列名字符串），所有标品共用
        scorecard_data:      手动传数组方式（备用）
        """
        name_map = {"train": "训练集", "test": "验证集", "oot": "OOT"}
        datasets = {}
        month_col_data = {} if month_col else None
        gain_score_data = {} if (gain_score_col and gain_label_col) else None

        for key in ["train", "test", "oot"]:
            sub = df[df[dataset_col] == key].reset_index(drop=True)
            if sub.empty:
                continue
            name = name_map[key]
            w = sub[weight_col].values if weight_col else None
            datasets[name] = (sub[target_col].values, sub[pred_col].values) + ((w,) if w is not None else ())
            if month_col and month_col in sub.columns:
                month_col_data[name] = sub[[month_col]].reset_index(drop=True)
            if gain_score_data is not None and gain_score_col in sub.columns and gain_label_col in sub.columns:
                gain_score_data[name] = (sub[gain_score_col].values, sub[gain_label_col].values)

        return self.classification_report(
            datasets,
            filename=filename,
            month_col_data=month_col_data if month_col_data else None,
            feature_importance=feature_importance,
            gain_score_data=gain_score_data if gain_score_data else None,
            scorecard_cols=scorecard_cols,
            scorecard_label_cols=scorecard_label_cols,
            scorecard_data=scorecard_data,
            raw_df=df,
            dataset_col=dataset_col,
        )

    # ------------------------------------------------------------------ #
    #  DataFrame 入口 — 回归
    # ------------------------------------------------------------------ #
    def regression_report_from_df(
        self,
        df: pd.DataFrame,
        target_col: str,
        pred_col: str,
        dataset_col: str = "dataset",
        month_col: Optional[str] = None,
        label_col: Optional[str] = None,
        gain_score_col: Optional[str] = None,
        gain_label_col: Optional[str] = None,
        scorecard_cols: Optional[list] = None,
        scorecard_label_cols=None,
        scorecard_data: Optional[dict] = None,
        filename: str = "回归模型评估报告",
        n_bins: int = 10,
        bins: list = None,
        n_bins_pred: int = 10,
        bins_pred: list = None,
        feature_importance: Optional[pd.DataFrame] = None,
    ) -> dict:
        """
        直接传入完整 DataFrame，内部按 dataset_col 拆分为 train / test / oot 三个集合。

        dataset_col 的值约定：'train' = 训练集，'test' = 验证集，'oot' = OOT
        month_col:           月份列名，传入后生成按月分桶和按月评估 sheet
        label_col:           二值列名（0/1），传入后在分桶表和矩阵中加 1值占比
        gain_score_col:      score 列名，与 gain_label_col 配合生成增益矩阵 sheet
        gain_label_col:      逾期标签列名（0/1），与 gain_score_col 配合使用
        scorecard_cols:      标品分列名列表，与 scorecard_label_col 配合生成标品分析 sheet
        scorecard_label_cols: 逾期标签列名列表（或单个列名字符串），所有标品共用
        scorecard_data:      手动传数组方式（备用）
        切点由训练集（train）计算，三集合共享
        """
        name_map = {"train": "训练集", "test": "验证集", "oot": "OOT"}
        datasets = {}
        month_col_data = {} if month_col else None
        label_col_data = {} if label_col else None

        for key in ["train", "test", "oot"]:
            sub = df[df[dataset_col] == key].reset_index(drop=True)
            if sub.empty:
                continue
            name = name_map[key]
            datasets[name] = (sub[target_col].values, sub[pred_col].values)
            if month_col and month_col in sub.columns:
                month_col_data[name] = sub[[month_col]].reset_index(drop=True)
            if label_col and label_col in sub.columns:
                label_col_data[name] = sub[label_col].values

        return self.regression_report(
            datasets,
            filename=filename,
            n_bins=n_bins,
            bins=bins,
            n_bins_pred=n_bins_pred,
            bins_pred=bins_pred,
            month_col_data=month_col_data if month_col_data else None,
            label_col_data=label_col_data if label_col_data else None,
            feature_importance=feature_importance,
            raw_df=df,
            target_col=target_col,
            gain_score_col=gain_score_col,
            gain_label_col=gain_label_col,
            scorecard_cols=scorecard_cols,
            scorecard_label_cols=scorecard_label_cols,
            scorecard_data=scorecard_data,
        )

    # ------------------------------------------------------------------ #
    #  特征分析报告（整体 + by dataset + by month，汇总到一个 Excel）
    # ------------------------------------------------------------------ #
    def feature_analysis_report(
        self,
        df: pd.DataFrame,
        features: list,
        dataset_col: Optional[str] = "dataset",
        month_col: Optional[str] = "month",
        cat_cols: Optional[list] = None,
        filename: str = "特征分析报告",
        target_col: Optional[str] = None,
    ) -> str:
        """
        将整体、by dataset、by month 三个维度的缺失率 / 一值率 / 分位数统计
        汇总输出到同一个 Excel，共 3 个 sheet：
          - 整体        : 缺失率 → 空行 → 一值率 → 空行 → 分位数统计
          - by_dataset  : 同上，各指标前有分组列
          - by_month    : 同上，各指标前有分组列
        """
        from .feature_analysis import FeatureAnalyzer

        cat_cols = cat_cols or []
        analyzer = FeatureAnalyzer(df[features], cat_cols=cat_cols)

        def _merge_stats(mr, sr, qs) -> pd.DataFrame:
            """把缺失率、一值率、分位数横向合并为每特征一行。"""
            mr2 = mr.rename(columns={"缺失数": "缺失数", "缺失率": "缺失率"})
            sr2 = sr.rename(columns={"最高频值": "最高频值", "一值率": "一值率"})
            qs2 = qs.drop(columns=["样本数"], errors="ignore") if "样本数" in qs.columns else qs.copy()
            merged = mr2.merge(sr2, on="特征名", how="outer").merge(qs2, on="特征名", how="outer")
            return merged

        def _merge_group_stats(group_dict) -> pd.DataFrame:
            """把 by_group 结果的缺失率、一值率、分位数横向合并，每特征+分组一行。"""
            mr = group_dict.get("缺失率")
            sr = group_dict.get("一值率")
            qs = group_dict.get("分位数统计")
            if mr is None:
                return pd.DataFrame()
            # 分组列是第一列
            group_col = mr.columns[0]
            sr2 = sr.rename(columns={"最高频值": "最高频值", "一值率": "一值率"}) if sr is not None else None
            qs2 = qs.drop(columns=["样本数"], errors="ignore") if qs is not None and "样本数" in qs.columns else (qs.copy() if qs is not None else None)
            merged = mr
            if sr2 is not None:
                merged = merged.merge(sr2, on=[group_col, "特征名"], how="outer")
            if qs2 is not None:
                merged = merged.merge(qs2, on=[group_col, "特征名"], how="outer")
            return merged

        excel_path = self._path("01_feature_analysis", f"{filename}.xlsx")
        with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:

            # ── 整体 ──────────────────────────────────────────────────
            mr = analyzer.missing_rate()
            sr = analyzer.single_value_rate()
            qs = analyzer.quantile_stats()
            overall = _merge_stats(mr, sr, qs)
            if not overall.empty:
                overall.to_excel(writer, sheet_name="整体", index=False)

            # ── by dataset ────────────────────────────────────────────
            if dataset_col and dataset_col in df.columns:
                raw = df[features + [dataset_col]].copy()
                by_ds = analyzer.by_group_analysis(dataset_col, raw)
                ds_sheet = _merge_group_stats(by_ds)
                if not ds_sheet.empty:
                    ds_sheet.to_excel(writer, sheet_name="by_dataset", index=False)

            # ── by month ──────────────────────────────────────────────
            if month_col and month_col in df.columns:
                raw_m = df[features + [month_col]].copy()
                by_m = analyzer.by_group_analysis(month_col, raw_m)
                m_sheet = _merge_group_stats(by_m)
                if not m_sheet.empty:
                    m_sheet.to_excel(writer, sheet_name="by_month", index=False)

            # ── 特征分箱图（按 target_col 均值）─────────────────────────
            if (
                target_col is not None
                and dataset_col is not None
                and dataset_col in df.columns
            ):
                import matplotlib
                matplotlib.use("Agg")
                import matplotlib.pyplot as plt
                import matplotlib.font_manager as _fma
                import numpy as _np
                from io import BytesIO
                from openpyxl.drawing.image import Image as XLImage

                # 中文字体（Linux/Windows 兼容）
                _CN_CANDS = [
                    "Microsoft YaHei", "SimHei", "SimSun",
                    "Noto Sans SC", "Noto Sans CJK SC", "Noto Sans CJK JP",
                    "WenQuanYi Micro Hei", "WenQuanYi Zen Hei",
                    "AR PL UMing CN", "DejaVu Sans",
                ]
                _avail = {f.name for f in _fma.fontManager.ttflist}
                _cn_font = next((fn for fn in _CN_CANDS if fn in _avail), None)
                if _cn_font:
                    plt.rcParams["font.family"] = _cn_font
                plt.rcParams["axes.unicode_minus"] = False

                dataset_map = {"train": "Train", "test": "Test", "oot": "OOT"}
                present_datasets = [
                    d for d in ["train", "test", "oot"]
                    if d in df[dataset_col].unique()
                ]

                wb = writer.book
                ws = wb.create_sheet("特征分箱图")

                IMG_W = 480
                IMG_H = 320
                row_height    = int(IMG_H / 0.75 / 20) + 1  # Excel 行高单位
                col_width_img = 1                             # 每张图占1列，列宽单独设置
                name_col      = 1
                img_start_col = 2
                img_col_width = IMG_W / 7.0                   # 图片宽度换算为 Excel 列宽字符数

                train_df_fa = df[df[dataset_col] == "train"]

                for feat_idx, feat in enumerate(features):
                    try:
                        train_vals_fa = train_df_fa[feat].dropna()
                        if train_vals_fa.empty:
                            continue
                        quantiles_fa = _np.linspace(0, 100, 6)
                        cut_edges_fa = _np.unique(_np.percentile(train_vals_fa, quantiles_fa))
                        if len(cut_edges_fa) < 2:
                            continue
                        cut_edges_fa[0]  = -_np.inf
                        cut_edges_fa[-1] =  _np.inf
                        binned = pd.cut(df[feat], bins=cut_edges_fa, include_lowest=True, duplicates="drop")
                    except Exception:
                        continue

                    bin_labels = [str(b) for b in binned.cat.categories]
                    df_work = df[[feat, dataset_col, target_col]].copy()
                    df_work["__bin__"] = binned.values

                    anchor_row = feat_idx * row_height + 1
                    ws.cell(row=anchor_row, column=name_col, value=feat)

                    # 预计算全局 y 轴范围，保证同一特征三图可直接对比
                    all_means_g = []
                    ds_agg_cache_fa = {}
                    for ds_key in present_datasets:
                        subset_all = df_work[df_work[dataset_col] == ds_key].copy()
                        subset     = subset_all.dropna(subset=[feat])
                        total_ds   = len(subset)
                        agg_fa = subset.groupby("__bin__", observed=True).agg(
                            mean_val=(target_col, "mean"),
                            count_val=(target_col, "count"),
                        ).reindex(binned.cat.categories)
                        means_fa  = agg_fa["mean_val"].values
                        counts_fa = agg_fa["count_val"].fillna(0).values
                        pcts_fa   = counts_fa / len(subset_all) if len(subset_all) > 0 else counts_fa * 0

                        # Missing 箱
                        sub_miss_fa  = subset_all[subset_all[feat].isna()]
                        miss_mean_fa = float(sub_miss_fa[target_col].mean()) if not sub_miss_fa.empty else _np.nan
                        miss_n_fa    = len(sub_miss_fa)
                        miss_pct_fa  = miss_n_fa / len(subset_all) if len(subset_all) > 0 else 0.0

                        # 整体均值（含 Missing 样本，用于伪Lift）
                        ds_overall_mean_fa = float(subset_all[target_col].mean()) if not subset_all.empty else _np.nan

                        ds_agg_cache_fa[ds_key] = (means_fa, counts_fa, pcts_fa,
                                                   miss_mean_fa, miss_n_fa, miss_pct_fa,
                                                   ds_overall_mean_fa)
                        all_means_g.extend([v for v in means_fa if not _np.isnan(v)])
                        if not _np.isnan(miss_mean_fa):
                            all_means_g.append(miss_mean_fa)

                    if not all_means_g:
                        continue
                    g_ymin_fa = float(_np.min(all_means_g))
                    g_ymax_fa = float(_np.max(all_means_g))
                    g_pad_fa  = (g_ymax_fa - g_ymin_fa) * 0.35 if (g_ymax_fa - g_ymin_fa) > 0 else abs(g_ymax_fa) * 0.35 + 0.05
                    y_lo_fa   = g_ymin_fa - g_pad_fa * 0.15
                    y_hi_fa   = g_ymax_fa + g_pad_fa

                    for ds_idx, ds_key in enumerate(present_datasets):
                        ds_label = dataset_map[ds_key]
                        means, counts, pcts, miss_mean_fa, miss_n_fa, miss_pct_fa, ds_overall_mean_fa = ds_agg_cache_fa[ds_key]

                        has_miss_fa = miss_n_fa > 0 and not _np.isnan(miss_mean_fa)
                        all_bin_lbls = bin_labels + (["Missing"] if has_miss_fa else [])
                        all_means_p  = _np.append(means, miss_mean_fa) if has_miss_fa else means.copy()
                        all_pcts_p   = _np.append(pcts,  miss_pct_fa)  if has_miss_fa else pcts.copy()
                        x = _np.arange(len(all_bin_lbls))

                        fig, ax_bar = plt.subplots(figsize=(IMG_W / 96, IMG_H / 96))
                        ax_line = ax_bar.twinx()

                        # 柱子：高度 = 样本占比，左轴
                        bar_colors_fa = ["#4C72B0"] * len(means) + (["#B0784C"] if has_miss_fa else [])
                        ax_bar.bar(x, all_pcts_p, color=bar_colors_fa, edgecolor="white",
                                   alpha=0.75, zorder=2)
                        ax_bar.set_ylim(0, max(all_pcts_p.max() * 2.2, 0.1))
                        ax_bar.set_ylabel("Pct", fontsize=6, color="#4C72B0")
                        ax_bar.tick_params(axis="y", labelsize=5, colors="#4C72B0")

                        # 折线：高度 = 均值，右轴，全局统一范围
                        valid = ~_np.isnan(means)
                        vx = _np.arange(len(means))
                        if valid.sum() > 1:
                            ax_line.plot(vx[valid], means[valid], color="#E84646",
                                         linewidth=1.2, marker="o", markersize=3, zorder=3)
                        if has_miss_fa and not _np.isnan(miss_mean_fa):
                            ax_line.plot(len(means), miss_mean_fa,
                                         marker="o", markersize=3, color="#B0784C", alpha=0.5, zorder=3)
                        ax_line.set_ylim(y_lo_fa, y_hi_fa)
                        ax_line.set_ylabel("Mean", fontsize=6, color="#E84646")
                        ax_line.tick_params(axis="y", labelsize=5, colors="#E84646")

                        # 占比标注：柱顶上方，不依赖柱高
                        pct_pad_fa = ax_bar.get_ylim()[1] * 0.02
                        for xi, pv in enumerate(all_pcts_p):
                            if _np.isnan(pv) or all_pcts_p[xi] == 0:
                                continue
                            ax_bar.text(xi, all_pcts_p[xi] + pct_pad_fa, f"{pv:.1%}",
                                        ha="center", va="bottom", fontsize=5,
                                        color="#333333", zorder=4)

                        # 整体均值参考线（右轴坐标）
                        if not _np.isnan(ds_overall_mean_fa):
                            ax_line.axhline(ds_overall_mean_fa, color="#E84646", linewidth=0.8,
                                            linestyle="--", alpha=0.6, zorder=1)

                        # 均值标注：折线点正下方，边缘箱向内
                        mean_pad_fa = (y_hi_fa - y_lo_fa) * 0.04
                        n_bars_fa = len(all_means_p)
                        for xi, mv in enumerate(all_means_p):
                            if _np.isnan(mv):
                                continue
                            if xi >= n_bars_fa - 2:
                                ax_line.text(xi - 0.15, mv - mean_pad_fa, f"{mv:.3f}",
                                             ha="right", va="top", fontsize=4.5,
                                             color="#CC2222", zorder=4)
                            else:
                                ax_line.text(xi + 0.15, mv - mean_pad_fa, f"{mv:.3f}",
                                             ha="left", va="top", fontsize=4.5,
                                             color="#CC2222", zorder=4)

                        ax_bar.set_xticks(x)
                        ax_bar.set_xticklabels(all_bin_lbls, rotation=45, ha="right", fontsize=5.5)
                        ax_bar.set_title(f"{feat} - {ds_label}", fontsize=7, pad=3)
                        ax_bar.margins(x=0.05)
                        fig.subplots_adjust(left=0.12, right=0.88, top=0.88, bottom=0.28)

                        buf = BytesIO()
                        fig.savefig(buf, format="png", dpi=110)
                        plt.close(fig)
                        buf.seek(0)

                        img = XLImage(buf)
                        img.width  = IMG_W
                        img.height = IMG_H

                        anchor_col = img_start_col + ds_idx * col_width_img
                        ws.add_image(img, ws.cell(row=anchor_row, column=anchor_col).coordinate)

                # 设置列宽：特征名列窄，图片列按图宽设置
                from openpyxl.utils import get_column_letter
                ws.column_dimensions[get_column_letter(name_col)].width = 15
                for ds_idx in range(len(present_datasets)):
                    c = img_start_col + ds_idx * col_width_img
                    ws.column_dimensions[get_column_letter(c)].width = img_col_width
                # 设置每行行高
                for feat_idx in range(len(features)):
                    anchor_row = feat_idx * row_height + 1
                    ws.row_dimensions[anchor_row].height = IMG_H / 0.75

        return excel_path

    # ------------------------------------------------------------------ #
    #  特征筛选 Excel 报告
    # ------------------------------------------------------------------ #
    def feature_selection_report(
        self,
        feature_report_df: pd.DataFrame,
        analysis_results: Optional[dict] = None,
        filename: str = "特征筛选报告",
    ) -> str:
        excel_path = self._path("02_feature_selection", f"{filename}.xlsx")
        with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
            feature_report_df.to_excel(writer, sheet_name="特征筛选结果", index=False)
            if analysis_results:
                for sheet_name, df in analysis_results.items():
                    if isinstance(df, pd.DataFrame) and not df.empty:
                        df.to_excel(writer, sheet_name=sheet_name[:31], index=False)
        return excel_path

    # ------------------------------------------------------------------ #
    #  特征分箱报告（按 dataset 分别输出）
    # ------------------------------------------------------------------ #
    def feature_bin_report(
        self,
        df: pd.DataFrame,
        target: str,
        features: list,
        dataset_col: str = "dataset",
        binning=None,
        n_bins: int = 10,
        filename: str = "特征分箱报告",
    ) -> str:
        """
        按 dataset 列（train/test/oot）分别输出每个特征的分箱统计：
        各箱坏样本占比、样本占比

        参数
        ----
        df          : 含特征列、target 列、dataset_col 列的完整 DataFrame
        target      : 二分类标签列名
        features    : 需要输出分箱图的特征列表
        dataset_col : 区分 train/test/oot 的列名，默认 'dataset'
        binning     : 已拟合的 Binning 对象；若为 None，则用训练集重新拟合
        n_bins      : binning 为 None 时使用的分箱数
        filename    : 输出文件名（不含扩展名）
        """
        from .binning import Binning

        # 用训练集确定箱边界
        if binning is None:
            train_df = df[df[dataset_col] == "train"].reset_index(drop=True)
            if train_df.empty:
                train_df = df.reset_index(drop=True)
            binning = Binning(n_bins=n_bins)
            binning.fit(train_df[features + [target]], target=target, cols=features)

        datasets = sorted(df[dataset_col].dropna().unique())
        excel_path = self._path("01_feature_analysis", f"{filename}.xlsx")

        with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
            for feat in features:
                if feat not in binning._bin_edges:
                    continue
                edges = binning._bin_edges[feat]
                all_rows = []
                for ds in datasets:
                    sub = df[df[dataset_col] == ds][[feat, target]].copy()
                    if sub.empty:
                        continue
                    total_n   = len(sub)
                    total_bad = int(sub[target].sum())

                    # 正常箱
                    sub_valid = sub.dropna(subset=[feat]).copy()
                    sub_valid["_bin"] = pd.cut(sub_valid[feat], bins=edges, include_lowest=True)
                    grp_rows = []
                    for bin_label, g in sub_valid.groupby("_bin"):
                        n   = len(g)
                        bad = int(g[target].sum())
                        grp_rows.append({
                            "dataset":     ds,
                            "分箱区间":    str(bin_label),
                            "样本数":      n,
                            "样本占比":    round(n / total_n, 4) if total_n > 0 else None,
                            "坏样本数":    bad,
                            "坏样本率":    round(bad / n, 4) if n > 0 else None,
                            "坏样本在全集占比": round(bad / total_bad, 4) if total_bad > 0 else None,
                        })

                    # Missing 箱（特征值为空）
                    sub_miss = sub[sub[feat].isna()]
                    if not sub_miss.empty:
                        mn   = len(sub_miss)
                        mbad = int(sub_miss[target].sum())
                        grp_rows.append({
                            "dataset":     ds,
                            "分箱区间":    "Missing",
                            "样本数":      mn,
                            "样本占比":    round(mn / total_n, 4) if total_n > 0 else None,
                            "坏样本数":    mbad,
                            "坏样本率":    round(mbad / mn, 4) if mn > 0 else None,
                            "坏样本在全集占比": round(mbad / total_bad, 4) if total_bad > 0 else None,
                        })

                    # All 合计行（含 Missing 样本）
                    grp_rows.append({
                        "dataset":     ds,
                        "分箱区间":    "All",
                        "样本数":      total_n,
                        "样本占比":    1.0,
                        "坏样本数":    total_bad,
                        "坏样本率":    round(total_bad / total_n, 4) if total_n > 0 else None,
                        "坏样本在全集占比": 1.0 if total_bad > 0 else None,
                    })

                    all_rows.extend(grp_rows)

                if all_rows:
                    feat_df = pd.DataFrame(all_rows)
                    sheet = feat[:31]
                    feat_df.to_excel(writer, sheet_name=sheet, index=False)

        return excel_path

    # ------------------------------------------------------------------ #
    #  内部辅助
    # ------------------------------------------------------------------ #
    @staticmethod
    def _df_to_html_table(df: pd.DataFrame) -> str:
        return df.to_html(index=False, border=1, classes="tbl", justify="center")

    def _build_clf_html(self, summary_df, sheets, title) -> str:
        css = ("<style>body{font-family:微软雅黑,Arial;margin:20px}"
               "h1,h2{color:#2c3e50}.tbl{border-collapse:collapse;width:100%;margin-bottom:20px}"
               ".tbl th{background:#2c3e50;color:#fff;padding:6px 10px}"
               ".tbl td{padding:5px 10px;border:1px solid #ccc}"
               ".tbl tr:nth-child(even){background:#f2f2f2}</style>")
        body = f"<h1>{title}</h1><h2>汇总指标</h2>" + self._df_to_html_table(summary_df)
        for name, df in sheets.items():
            body += f"<h2>{name}</h2>" + self._df_to_html_table(df)
        return (f"<!DOCTYPE html><html><head><meta charset='utf-8'>"
                f"<title>{title}</title>{css}</head><body>{body}</body></html>")

    def _build_reg_html(self, summary_df, sheets, title) -> str:
        css = ("<style>body{font-family:微软雅黑,Arial;margin:20px}"
               "h1,h2{color:#2c3e50}.tbl{border-collapse:collapse;width:100%;margin-bottom:20px}"
               ".tbl th{background:#2c3e50;color:#fff;padding:6px 10px}"
               ".tbl td{padding:5px 10px;border:1px solid #ccc}"
               ".tbl tr:nth-child(even){background:#f2f2f2}</style>")
        body = f"<h1>{title}</h1><h2>汇总指标</h2>" + self._df_to_html_table(summary_df)
        for name, df in sheets.items():
            body += f"<h2>{name}</h2>" + self._df_to_html_table(df)
        return (f"<!DOCTYPE html><html><head><meta charset='utf-8'>"
                f"<title>{title}</title>{css}</head><body>{body}</body></html>")


# ------------------------------------------------------------------ #
#  模块级工具函数
# ------------------------------------------------------------------ #
def _detect_month_col(df: pd.DataFrame) -> Optional[str]:
    """自动检测月份列名（优先 'month'，其次 'ym'，再次含 'month'/'ym' 的列）"""
    for cand in ["month", "ym", "月份", "yearmonth", "year_month"]:
        if cand in df.columns:
            return cand
    for col in df.columns:
        if "month" in col.lower() or "ym" in col.lower():
            return col
    return None
