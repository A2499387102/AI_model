import os
import json
import joblib
import pandas as pd
import numpy as np
from typing import Optional
from sklearn.metrics import roc_auc_score, roc_curve, mean_squared_error


def _ks_auc(y_true, y_score, weights=None):
    fpr, tpr, _ = roc_curve(y_true, y_score, sample_weight=weights)
    auc = roc_auc_score(y_true, y_score, sample_weight=weights)
    if auc < 0.5:
        auc = 1 - auc
    ks = round(float(np.abs(tpr - fpr).max()), 4)
    return ks, round(float(auc), 4)


def _run_hyperopt(objective, space, n_trials: int, desc: str):
    """
    封装 fmin + tqdm 进度条。
    每轮 objective 完成后自动更新进度条，后缀显示当前 best_loss。
    返回 Trials 对象（与直接调用 fmin 一致）。
    """
    from hyperopt import fmin, tpe, Trials, STATUS_OK  # noqa
    try:
        from tqdm.auto import tqdm
    except ImportError:
        tqdm = None

    trials = Trials()
    best_loss = [float("inf")]

    if tqdm is not None:
        pbar = tqdm(total=n_trials, desc=desc, ncols=90, leave=True)

        def wrapped(params):
            result = objective(params)
            best_loss[0] = min(best_loss[0], result["loss"])
            pbar.update(1)
            pbar.set_postfix({"best_loss": f"{best_loss[0]:.5f}"})
            return result

        try:
            fmin(fn=wrapped, space=space, algo=tpe.suggest,
                 max_evals=n_trials, trials=trials, verbose=False)
        finally:
            pbar.close()
    else:
        fmin(fn=objective, space=space, algo=tpe.suggest,
             max_evals=n_trials, trials=trials, verbose=False)

    return trials


def _build_trials_df(logs1: list, logs2: list) -> pd.DataFrame:
    """
    把两轮的 log 列表合并成一个 flat DataFrame：
      - params 字典展开成独立列（带 p_ 前缀，避免与指标列撞名）
      - 添加 round（第1轮/第2轮）、trial（轮内序号，从1开始）、is_best 列
    """
    def _flatten(records, round_label):
        rows = []
        for i, rec in enumerate(records):
            row = {"round": round_label, "trial": i + 1}
            # 指标列（非 params 的所有 key）
            for k, v in rec.items():
                if k != "params":
                    row[k] = v
            # 参数列展开
            params = rec.get("params", {})
            for k, v in params.items():
                row[f"p_{k}"] = v
            rows.append(row)
        return rows

    rows = _flatten(logs1, "第1轮") + _flatten(logs2, "第2轮")
    df = pd.DataFrame(rows)

    # is_best：整体 loss 最小的行标为 True
    if "loss" in df.columns and not df.empty:
        df["is_best"] = df["loss"] == df["loss"].min()
    else:
        df["is_best"] = False

    # 列排序：round/trial/is_best 放最前，指标列次之，参数列最后
    front = [c for c in ["round", "trial", "loss", "is_best"] if c in df.columns]
    metric_cols = [c for c in df.columns if c not in front and not c.startswith("p_")]
    param_cols  = [c for c in df.columns if c.startswith("p_")]
    df = df[front + metric_cols + param_cols]
    return df


def _save_tuning_log(df: pd.DataFrame, save_dir: str, filename: str = "tuning_log.xlsx") -> None:
    """把试验日志 DataFrame 写入 save_dir/03_model_tuning/tuning_log.xlsx，is_best 行高亮绿色。"""
    try:
        from openpyxl.styles import PatternFill
        tuning_dir = os.path.join(save_dir, "03_model_tuning")
        os.makedirs(tuning_dir, exist_ok=True)
        path = os.path.join(tuning_dir, filename)
        with pd.ExcelWriter(path, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name="调参日志", index=False)
            ws = writer.sheets["调参日志"]
            green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            is_best_col = df.columns.tolist().index("is_best") + 1  # 1-indexed
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                if row[is_best_col - 1].value is True:
                    for cell in row:
                        cell.fill = green
        print(f"调参日志已保存至: {path}")
    except Exception as e:
        print(f"[警告] tuning_log.xlsx 写入失败: {e}")


class LGBMTrainer:
    """LightGBM 二分类训练器"""

    def __init__(
        self,
        n_trials: int = 150,
        use_gpu: bool = False,
        cat_cols: Optional[list] = None,
        random_state: int = 2024,
    ):
        self.n_trials = n_trials
        self.use_gpu = use_gpu
        self.cat_cols = cat_cols or []
        self.random_state = random_state
        self.model = None
        self.best_params = None
        self.selected_features: list = []
        self.trials_log: pd.DataFrame = pd.DataFrame()

    def get_default_params(self, n_data: int) -> dict:
        """
        返回第一轮固定训练参数（全部为固定值，不含 hp 分布）。
        可通过 fit() 的 custom_params 参数覆盖任意 key。
        """
        return {
            "boosting_type": "gbdt", "objective": "binary", "metric": "auc",
            "is_unbalance": True, "nthread": -1, "verbose": -1, "bagging_freq": 2,
            "num_leaves": 64, "max_depth": 6, "n_estimators": 500,
            "learning_rate": 0.05, "bagging_fraction": 0.8, "feature_fraction": 0.8,
            "lambda_l1": 5, "lambda_l2": 5,
            "min_data_in_leaf": max(int(n_data * 0.01), 1),
        }

    def get_default_space(self, n_data: int) -> dict:
        """
        返回第二轮 Hyperopt 搜索空间字典。

        在 Jupyter 中不重启修改参数的用法：
            space = trainer.get_default_space(len(X_train))
            # 修改想调整的参数范围，例如：
            from hyperopt import hp
            space['learning_rate'] = hp.uniform('learning_rate', 0.001, 0.05)
            space['max_depth'] = hp.quniform('max_depth', 3, 6, 1)
            # 传入 fit
            trainer.fit(..., custom_space=space)
        """
        from hyperopt import hp
        return {
            "boosting_type": "gbdt", "objective": "binary", "metric": "auc",
            "is_unbalance": True, "feature_pre_filter": True, "nthread": -1, "verbose": -1,
            "bagging_freq": 2,
            "bagging_fraction": hp.quniform("bagging_fraction", 0.5, 0.8, 0.1),
            "feature_fraction": hp.quniform("feature_fraction", 0.5, 0.8, 0.1),
            "max_depth": hp.quniform("max_depth", 2, 4, 1),
            "num_leaves": hp.quniform("num_leaves", 2, 50, 2),
            "n_estimators": hp.quniform("n_estimators", 300, 500, 10),
            "learning_rate": hp.uniform("learning_rate", 0.005, 0.02),
            "lambda_l1": hp.randint("lambda_l1", 10, 50),
            "lambda_l2": hp.randint("lambda_l2", 10, 50),
            "min_data_in_leaf": hp.quniform(
                "min_data_in_leaf",
                max(int(n_data * 0.01), 1),
                max(int(n_data * 0.05), 2),
                100,
            ),
        }

    def fit(
        self,
        X_train: pd.DataFrame,
        y_train: pd.Series,
        X_test: pd.DataFrame,
        y_test: pd.Series,
        X_oot: pd.DataFrame,
        y_oot: pd.Series,
        train_weight: Optional[pd.Series] = None,
        test_weight: Optional[pd.Series] = None,
        oot_weight: Optional[pd.Series] = None,
        features: Optional[list] = None,
        custom_params: Optional[dict] = None,
        custom_space: Optional[dict] = None,
        save_dir: Optional[str] = None,
    ) -> "LGBMTrainer":
        try:
            import lightgbm as lgb
            from hyperopt import fmin, tpe, hp, Trials, STATUS_OK
        except ImportError as e:
            raise ImportError(f"请安装依赖: pip install lightgbm hyperopt — {e}")

        # 兼容 LightGBM 2.x（early_stopping_rounds/verbose_eval 直接参数）
        # 和 3.3+（callbacks API）
        _lgb_ver = tuple(int(x) for x in lgb.__version__.split(".")[:2])
        if _lgb_ver >= (3, 3):
            def _train(params, train_set, num_boost_round, valid_sets):
                return lgb.train(params, train_set, num_boost_round=num_boost_round,
                                 valid_sets=valid_sets,
                                 callbacks=[lgb.early_stopping(50, verbose=False),
                                            lgb.log_evaluation(-1)])
        else:
            def _train(params, train_set, num_boost_round, valid_sets):
                return lgb.train(params, train_set, num_boost_round=num_boost_round,
                                 valid_sets=valid_sets,
                                 early_stopping_rounds=50, verbose_eval=False)

        features = features or X_train.columns.tolist()
        Xtr  = X_train[features].reset_index(drop=True)
        Xts  = X_test[features].reset_index(drop=True)
        Xoot = X_oot[features].reset_index(drop=True)
        y_train = y_train.reset_index(drop=True)
        y_test  = y_test.reset_index(drop=True)
        y_oot   = y_oot.reset_index(drop=True)
        tw = train_weight.reset_index(drop=True) if train_weight is not None else pd.Series(np.ones(len(Xtr)))
        vw = test_weight.reset_index(drop=True)  if test_weight  is not None else pd.Series(np.ones(len(Xts)))
        ow = oot_weight.reset_index(drop=True)   if oot_weight   is not None else pd.Series(np.ones(len(Xoot)))

        n_data    = len(Xtr)
        train_set = lgb.Dataset(Xtr, y_train, categorical_feature=self.cat_cols, weight=tw, free_raw_data=False)
        test_set  = lgb.Dataset(Xts, y_test,  categorical_feature=self.cat_cols, weight=vw, free_raw_data=False)

        # ── 第一轮：固定参数单次训练，用于特征筛选 ────────────────────────────
        print("LGBM分类 第1轮（固定参数训练）")
        p1_fixed = {**self.get_default_params(n_data), **(custom_params or {})}
        num_br1  = int(p1_fixed.pop("n_estimators", 500))
        model1   = _train(p1_fixed, train_set, num_br1, test_set)
        tr_auc1  = roc_auc_score(y_train, model1.predict(Xtr,  num_iteration=model1.best_iteration), sample_weight=tw)
        ts_auc1  = roc_auc_score(y_test,  model1.predict(Xts,  num_iteration=model1.best_iteration), sample_weight=vw)
        oot_auc1 = roc_auc_score(y_oot,   model1.predict(Xoot, num_iteration=model1.best_iteration), sample_weight=ow)
        loss1    = -oot_auc1 + abs(tr_auc1 - ts_auc1)
        logs = [{"train_auc": round(tr_auc1, 4), "test_auc": round(ts_auc1, 4),
                 "oot_auc": round(oot_auc1, 4), "loss": round(loss1, 4),
                 "params": {**p1_fixed, "n_estimators": num_br1}}]

        imp = pd.DataFrame({"feature": model1.feature_name(), "importance": model1.feature_importance()})
        self.selected_features = imp[imp["importance"] > 0]["feature"].tolist()
        if not self.selected_features:
            self.selected_features = features

        Xtr2, Xts2, Xoot2 = Xtr[self.selected_features], Xts[self.selected_features], Xoot[self.selected_features]
        train_set2 = lgb.Dataset(Xtr2, y_train, categorical_feature=self.cat_cols, weight=tw, free_raw_data=False)
        test_set2  = lgb.Dataset(Xts2, y_test,  categorical_feature=self.cat_cols, weight=vw, free_raw_data=False)

        logs2 = []

        def objective2(params):
            params["num_leaves"]      = int(params["num_leaves"])
            params["max_depth"]       = int(params["max_depth"])
            num_boost_round2          = int(params.pop("n_estimators"))
            params["lambda_l1"]       = int(params["lambda_l1"])
            params["lambda_l2"]       = int(params["lambda_l2"])
            params["min_data_in_leaf"]= int(params["min_data_in_leaf"])
            m = _train(params, train_set2, num_boost_round2, test_set2)
            tr_auc  = roc_auc_score(y_train, m.predict(Xtr2,  num_iteration=m.best_iteration), sample_weight=tw)
            ts_auc  = roc_auc_score(y_test,  m.predict(Xts2,  num_iteration=m.best_iteration), sample_weight=vw)
            oot_auc = roc_auc_score(y_oot,   m.predict(Xoot2, num_iteration=m.best_iteration), sample_weight=ow)
            loss = -oot_auc + abs(tr_auc - ts_auc)
            logs2.append({"train_auc": round(tr_auc, 4), "test_auc": round(ts_auc, 4),
                          "oot_auc": round(oot_auc, 4), "loss": round(loss, 4),
                          "params": {**params, "n_estimators": num_boost_round2}})
            return {"loss": loss, "status": STATUS_OK}

        space2  = custom_space if custom_space is not None else self.get_default_space(n_data)
        trials2 = _run_hyperopt(objective2, space2, self.n_trials, "LGBM分类 第2轮")

        idx2    = int(np.argmin([r["loss"] for r in logs2]))
        params2 = trials2.trials[idx2]["misc"]["vals"]
        params2 = {k: v[0] if isinstance(v, list) and len(v) == 1 else v for k, v in params2.items()}
        num_br_final = int(params2.pop("n_estimators", 300))
        params2.update({"boosting_type": "gbdt", "objective": "binary", "metric": "auc",
                        "is_unbalance": True, "feature_pre_filter": False,
                        "nthread": -1, "verbose": -1, "bagging_freq": 2})
        for k in ["num_leaves", "max_depth", "lambda_l1", "lambda_l2", "min_data_in_leaf"]:
            if k in params2:
                params2[k] = int(params2[k])
        self.model = _train(params2, train_set2, num_br_final, test_set2)
        self.best_params = {**params2, "n_estimators": num_br_final}
        self.trials_log  = _build_trials_df(logs, logs2)

        if save_dir:
            self.save(save_dir)
        return self

    def predict_proba(self, X: pd.DataFrame) -> np.ndarray:
        feats = self.selected_features or X.columns.tolist()
        return self.model.predict(X[feats], num_iteration=self.model.best_iteration)

    def save(self, save_dir: str) -> None:
        """保存模型到 save_dir/05_model_deploy/model.pkl（含模型、入模特征、最优参数）"""
        deploy_dir = os.path.join(save_dir, "05_model_deploy")
        os.makedirs(deploy_dir, exist_ok=True)
        payload = {
            "model_type": "lgbm",
            "model": self.model,
            "selected_features": self.selected_features,
            "best_params": self.best_params,
        }
        joblib.dump(payload, os.path.join(deploy_dir, "model.pkl"))
        if not self.trials_log.empty:
            _save_tuning_log(self.trials_log, save_dir)
        print(f"模型已保存至: {deploy_dir}")

    @classmethod
    def load(cls, save_dir: str) -> "LGBMTrainer":
        """从 save_dir/model.pkl 恢复模型"""
        payload = joblib.load(os.path.join(save_dir, "05_model_deploy", "model.pkl"))
        trainer = cls()
        trainer.model = payload["model"]
        trainer.selected_features = payload["selected_features"]
        trainer.best_params = payload["best_params"]
        return trainer


class XGBTrainer:
    """XGBoost 二分类训练器"""

    def __init__(self, n_trials: int = 100, use_gpu: bool = False, random_state: int = 2024):
        self.n_trials = n_trials
        self.use_gpu = use_gpu
        self.random_state = random_state
        self.model = None
        self.best_params = None
        self.selected_features: list = []
        self.trials_log: pd.DataFrame = pd.DataFrame()

    def get_default_space(self, odds: float) -> dict:
        """
        返回默认超参数搜索空间字典。

        在 Jupyter 中不重启修改参数的用法：
            space = trainer.get_default_space(odds)
            from hyperopt import hp
            space['learning_rate'] = hp.quniform('learning_rate', 0.005, 0.1, 0.005)
            trainer.fit(..., custom_space=space)
        """
        from hyperopt import hp
        return {
            "base_score": 0.5, "booster": "gbtree", "objective": "binary:logistic",
            "learning_rate": hp.quniform("learning_rate", 0.01, 0.3, 0.01),
            "gamma": hp.quniform("gamma", 0, 200, 2),
            "max_depth": hp.choice("max_depth", [2, 3]),
            "subsample": hp.quniform("subsample", 0.5, 1, 0.1),
            "colsample_bytree": 1,
            "n_estimators": 500,
            "min_child_weight": hp.quniform("min_child_weight", 100, 1000, 50),
            "reg_alpha": hp.quniform("reg_alpha", 0, 300, 10),
            "reg_lambda": hp.quniform("reg_lambda", 0, 300, 10),
            "scale_pos_weight": hp.quniform("scale_pos_weight", max(int(odds) - 2, 1), int(odds) + 2, 1),
            "random_state": self.random_state,
            "tree_method": "gpu_hist" if self.use_gpu else "hist",
            "early_stopping_rounds": 30,
            "eval_metric": "auc",
            "nthread": -1,
        }

    @staticmethod
    def get_default_params(odds: float) -> dict:
        """返回第一轮固定训练参数（odds = 负正样本比）"""
        return {
            "base_score": 0.5, "booster": "gbtree", "objective": "binary:logistic",
            "eval_metric": "auc", "n_estimators": 500,
            "learning_rate": 0.05, "max_depth": 4,
            "subsample": 0.8, "colsample_bytree": 0.8,
            "min_child_weight": max(int(odds * 10), 10),
            "gamma": 5, "reg_alpha": 10, "reg_lambda": 10,
            "scale_pos_weight": max(int(odds), 1),
            "early_stopping_rounds": 30, "nthread": -1,
        }

    def fit(
        self,
        X_train: pd.DataFrame, y_train: pd.Series,
        X_test:  pd.DataFrame, y_test:  pd.Series,
        X_oot:   pd.DataFrame, y_oot:   pd.Series,
        train_weight: Optional[pd.Series] = None,
        test_weight:  Optional[pd.Series] = None,
        oot_weight:   Optional[pd.Series] = None,
        features: Optional[list] = None,
        custom_space: Optional[dict] = None,
        custom_params: Optional[dict] = None,
        save_dir: Optional[str] = None,
    ) -> "XGBTrainer":
        try:
            import xgboost as xgb
            from hyperopt import fmin, tpe, hp, Trials, STATUS_OK
        except ImportError as e:
            raise ImportError(f"请安装依赖: pip install xgboost hyperopt — {e}")

        features = features or X_train.columns.tolist()
        Xtr  = X_train[features].reset_index(drop=True)
        Xts  = X_test[features].reset_index(drop=True)
        Xoot = X_oot[features].reset_index(drop=True)
        y_train = y_train.reset_index(drop=True)
        y_test  = y_test.reset_index(drop=True)
        y_oot   = y_oot.reset_index(drop=True)
        tw = train_weight.reset_index(drop=True) if train_weight is not None else pd.Series(np.ones(len(Xtr)))
        vw = test_weight.reset_index(drop=True)  if test_weight  is not None else pd.Series(np.ones(len(Xts)))
        ow = oot_weight.reset_index(drop=True)   if oot_weight   is not None else pd.Series(np.ones(len(Xoot)))

        tree_method = "gpu_hist" if self.use_gpu else "hist"
        odds = max(float((tw[y_train == 0].sum()) / (tw[y_train == 1].sum() + 1e-6)), 1.0)

        # ── 第1轮：固定参数训练 ──────────────────────────────────────────────
        print("XGB分类 第1轮（固定参数训练）")
        p1 = self.get_default_params(odds)
        p1.update(custom_params or {})
        p1["tree_method"]  = tree_method
        p1["random_state"] = self.random_state
        model1 = xgb.XGBClassifier(**p1, verbosity=0)
        model1.fit(Xtr, y_train, sample_weight=tw,
                   eval_set=[(Xts, y_test)], sample_weight_eval_set=[vw], verbose=False)
        _, tr_auc1  = _ks_auc(y_train, model1.predict_proba(Xtr)[:, 1], tw)
        _, ts_auc1  = _ks_auc(y_test,  model1.predict_proba(Xts)[:, 1], vw)
        _, oot_auc1 = _ks_auc(y_oot,   model1.predict_proba(Xoot)[:, 1], ow)
        loss1 = round(-oot_auc1 + abs(tr_auc1 - oot_auc1), 4)
        logs1 = [{"train_auc": round(tr_auc1, 4), "test_auc": round(ts_auc1, 4),
                  "oot_auc": round(oot_auc1, 4), "loss": loss1}]

        # ── 特征筛选 ────────────────────────────────────────────────────────
        imp = {k: v for k, v in model1.get_booster().get_score(importance_type="gain").items() if v > 0}
        self.selected_features = list(imp.keys()) if imp else features
        Xtr2  = Xtr[self.selected_features]
        Xts2  = Xts[self.selected_features]
        Xoot2 = Xoot[self.selected_features]

        # ── 第2轮：Hyperopt 超参搜索 ─────────────────────────────────────────
        logs2 = []

        def objective2(params):
            clf = xgb.XGBClassifier(**params, verbosity=0)
            clf.fit(Xtr2, y_train, sample_weight=tw,
                    eval_set=[(Xts2, y_test)], sample_weight_eval_set=[vw], verbose=False)
            _, tr_auc  = _ks_auc(y_train, clf.predict_proba(Xtr2)[:, 1], tw)
            _, ts_auc  = _ks_auc(y_test,  clf.predict_proba(Xts2)[:, 1], vw)
            _, oot_auc = _ks_auc(y_oot,   clf.predict_proba(Xoot2)[:, 1], ow)
            loss = -oot_auc + abs(tr_auc - oot_auc)
            logs2.append({"loss": round(loss, 4), "train_auc": round(tr_auc, 4),
                          "test_auc": round(ts_auc, 4), "oot_auc": round(oot_auc, 4)})
            return {"loss": loss, "status": STATUS_OK}

        space2 = custom_space if custom_space is not None else self.get_default_space(odds)
        trials2 = _run_hyperopt(objective2, space2, self.n_trials, "XGB分类 第2轮")

        best_idx2 = int(np.argmin([r["loss"] for r in logs2]))
        best_t2   = trials2.trials[best_idx2]
        p2 = {k: v[0] if isinstance(v, list) else v for k, v in best_t2["misc"]["vals"].items()}
        p2.update({"base_score": 0.5, "booster": "gbtree", "objective": "binary:logistic",
                   "colsample_bytree": 1, "n_estimators": 1000, "random_state": self.random_state,
                   "tree_method": tree_method, "early_stopping_rounds": 30, "eval_metric": "auc", "nthread": -1})
        # max_depth 是 hp.choice，vals 存的是候选列表的 index，从 space2 取实际候选列表还原
        if "max_depth" in p2:
            from hyperopt.pyll import Apply
            md_space = space2.get("max_depth")
            if hasattr(md_space, "pos_args"):
                candidates = [a.obj for a in md_space.pos_args[1:]]
                p2["max_depth"] = candidates[int(p2["max_depth"])]
            else:
                p2["max_depth"] = int(p2["max_depth"])
        self.model = xgb.XGBClassifier(**p2, verbosity=0)
        self.model.fit(Xtr2, y_train, sample_weight=tw, eval_set=[(Xts2, y_test)],
                       sample_weight_eval_set=[vw], verbose=False)
        self.best_params = p2
        self.trials_log  = _build_trials_df(logs1, logs2)

        if save_dir:
            self.save(save_dir)
        return self

    def predict_proba(self, X: pd.DataFrame) -> np.ndarray:
        feats = self.selected_features or X.columns.tolist()
        return self.model.predict_proba(X[feats])[:, 1]

    def save(self, save_dir: str) -> None:
        deploy_dir = os.path.join(save_dir, "05_model_deploy")
        os.makedirs(deploy_dir, exist_ok=True)
        payload = {
            "model_type": "xgboost",
            "model": self.model,
            "selected_features": self.selected_features,
            "best_params": self.best_params,
        }
        joblib.dump(payload, os.path.join(deploy_dir, "model.pkl"))
        if not self.trials_log.empty:
            _save_tuning_log(self.trials_log, save_dir)
        print(f"模型已保存至: {deploy_dir}")

    @classmethod
    def load(cls, save_dir: str) -> "XGBTrainer":
        payload = joblib.load(os.path.join(save_dir, "05_model_deploy", "model.pkl"))
        trainer = cls()
        trainer.model = payload["model"]
        trainer.selected_features = payload["selected_features"]
        trainer.best_params = payload["best_params"]
        return trainer


def _rmse(y_true, y_pred, weights=None):
    if weights is not None:
        weights = np.array(weights)
        weights = weights / weights.sum()
        return round(float(np.sqrt(np.sum(weights * (np.array(y_true) - np.array(y_pred)) ** 2))), 4)
    return round(float(np.sqrt(mean_squared_error(y_true, y_pred))), 4)


def _reg_loss_mae(y_tr, p_tr, y_ts, p_ts, y_oo, p_oo, tw_, vw_, ow_, scheme="A"):
    """
    共用回归调参 loss（LGBM 回归 / XGBoost 回归均使用同一函数）。

    scheme='A'（默认）：
      loss = oot_mae
           + 0.5 * (oot_mae / tr_mae)           # OOT 相对训练集的衰退倍数（无量纲）
           + 0.3 * |ts_mae - oot_mae| / ts_mae  # test→OOT 时间漂移率（无量纲）

    scheme='B'：
      loss = oot_mae + 0.5 * |oot_mae - tr_mae|
      充分拟合 train（tr_mae 小）的前提下，惩罚 train 与 OOT 的绝对差距。
      alpha=0.5 可在构造 LGBMRegressorTrainer/XGBRegressorTrainer 时通过
      loss_scheme_alpha 参数调整。
    """
    from sklearn.metrics import mean_absolute_error
    tr_mae  = mean_absolute_error(y_tr, p_tr,  sample_weight=tw_)
    ts_mae  = mean_absolute_error(y_ts, p_ts,  sample_weight=vw_)
    oot_mae = mean_absolute_error(y_oo, p_oo,  sample_weight=ow_)
    if scheme == "B":
        loss = oot_mae + 0.5 * abs(oot_mae - tr_mae)
    else:
        oot_rel = oot_mae / (tr_mae + 1e-8)
        drift   = abs(ts_mae - oot_mae) / (ts_mae + 1e-8)
        loss    = oot_mae + 0.5 * oot_rel + 0.3 * drift
    return round(loss, 6), round(tr_mae, 4), round(ts_mae, 4), round(oot_mae, 4)


class LGBMRegressorTrainer:
    """LightGBM 回归训练器，固定参数第1轮筛特征 + Hyperopt 第2轮调参
    第1轮：get_default_params 固定参数单次训练，用 gain 重要性筛选特征。
    第2轮：在筛选后特征上做 Hyperopt 超参搜索。
    调参 loss 与 XGBRegressorTrainer 完全一致：
      oot_mae + 0.5*(oot_mae/tr_mae) + 0.3*|ts_mae-oot_mae|/ts_mae
    """

    def __init__(
        self,
        n_trials: int = 100,
        cat_cols: Optional[list] = None,
        random_state: int = 2024,
        loss_scheme: str = "A",
    ):
        self.n_trials = n_trials
        self.cat_cols = cat_cols or []
        self.random_state = random_state
        self.loss_scheme = loss_scheme  # "A"（默认）或 "B"（train-OOT 差距惩罚）
        self.model = None
        self.best_params = None
        self.selected_features: list = []
        self.trials_log: pd.DataFrame = pd.DataFrame()

    @staticmethod
    def get_default_params(n_data: int) -> dict:
        """第1轮固定训练参数（n_estimators 在 fit 中被 pop 出后传 num_boost_round）。"""
        return {
            "boosting_type": "gbdt",
            "objective": "regression",
            "metric": "mae",
            "num_threads": -1,
            "verbose": -1,
            "bagging_freq": 2,
            "num_leaves": 64,
            "max_depth": 6,
            "n_estimators": 800,
            "learning_rate": 0.05,
            "bagging_fraction": 0.8,
            "feature_fraction": 0.8,
            "lambda_l1": 5,
            "lambda_l2": 5,
            "min_data_in_leaf": max(int(n_data * 0.01), 1),
        }

    def get_default_space(self, n_data: int) -> dict:
        """
        返回默认超参数搜索空间字典。

        在 Jupyter 中不重启修改参数的用法：
            space = trainer.get_default_space(len(X_train))
            from hyperopt import hp
            space['learning_rate'] = hp.uniform('learning_rate', 0.001, 0.05)
            trainer.fit(..., custom_space=space)
        """
        from hyperopt import hp
        return {
            "boosting_type": "gbdt",
            "objective": "regression",
            "metric": "mae",
            "num_threads": -1,
            "verbose": -1,
            "bagging_freq": 2,
            "bagging_fraction": hp.quniform("bagging_fraction", 0.6, 1.0, 0.1),
            "feature_fraction": hp.quniform("feature_fraction", 0.6, 1.0, 0.1),
            "max_depth":        hp.quniform("max_depth", 4, 8, 1),
            "num_leaves":       hp.quniform("num_leaves", 16, 128, 8),
            "n_estimators":     hp.quniform("n_estimators", 500, 1000, 50),
            "learning_rate":    hp.uniform("learning_rate", 0.005, 0.05),
            "lambda_l1":        hp.randint("lambda_l1", 0, 20),
            "lambda_l2":        hp.randint("lambda_l2", 0, 20),
            "min_data_in_leaf": hp.quniform(
                "min_data_in_leaf",
                max(int(n_data * 0.005), 1),
                max(int(n_data * 0.02), 2),
                10,
            ),
        }

    def fit(
        self,
        X_train: pd.DataFrame, y_train: pd.Series,
        X_test:  pd.DataFrame, y_test:  pd.Series,
        X_oot:   pd.DataFrame, y_oot:   pd.Series,
        train_weight: Optional[pd.Series] = None,
        test_weight:  Optional[pd.Series] = None,
        oot_weight:   Optional[pd.Series] = None,
        features: Optional[list] = None,
        custom_space: Optional[dict] = None,
        custom_params: Optional[dict] = None,
        save_dir: Optional[str] = None,
    ) -> "LGBMRegressorTrainer":
        try:
            import lightgbm as lgb
            from hyperopt import fmin, tpe, hp, Trials, STATUS_OK
        except ImportError as e:
            raise ImportError(f"请安装依赖: pip install lightgbm hyperopt — {e}")

        # 兼容 LightGBM 2.x（early_stopping_rounds/verbose_eval 直接参数）
        # 和 3.3+（callbacks API）
        _lgb_ver = tuple(int(x) for x in lgb.__version__.split(".")[:2])
        if _lgb_ver >= (3, 3):
            def _train(params, train_set, num_boost_round, valid_sets):
                return lgb.train(params, train_set, num_boost_round=num_boost_round,
                                 valid_sets=valid_sets,
                                 callbacks=[lgb.early_stopping(50, verbose=False),
                                            lgb.log_evaluation(-1)])
        else:
            def _train(params, train_set, num_boost_round, valid_sets):
                return lgb.train(params, train_set, num_boost_round=num_boost_round,
                                 valid_sets=valid_sets,
                                 early_stopping_rounds=50, verbose_eval=False)

        features = features or X_train.columns.tolist()
        Xtr  = X_train[features].reset_index(drop=True)
        Xts  = X_test[features].reset_index(drop=True)
        Xoot = X_oot[features].reset_index(drop=True)
        y_train = y_train.reset_index(drop=True)
        y_test  = y_test.reset_index(drop=True)
        y_oot   = y_oot.reset_index(drop=True)
        tw = train_weight.reset_index(drop=True) if train_weight is not None else pd.Series(np.ones(len(Xtr)))
        vw = test_weight.reset_index(drop=True)  if test_weight  is not None else pd.Series(np.ones(len(Xts)))
        ow = oot_weight.reset_index(drop=True)   if oot_weight   is not None else pd.Series(np.ones(len(Xoot)))

        train_set = lgb.Dataset(Xtr, y_train, categorical_feature=self.cat_cols, weight=tw, free_raw_data=False)
        test_set  = lgb.Dataset(Xts, y_test,  categorical_feature=self.cat_cols, weight=vw, free_raw_data=False)

        n_data = len(Xtr)

        # ── 第1轮：固定参数单次训练，用于特征筛选 ────────────────────────────
        print("LGBM回归 第1轮（固定参数训练）")
        p1 = self.get_default_params(n_data)
        if custom_params:
            p1.update(custom_params)
        num_br1 = int(p1.pop("n_estimators"))
        model1 = _train(p1, train_set, num_br1, test_set)
        loss1, tr_mae1, ts_mae1, oot_mae1 = _reg_loss_mae(
            y_train.values, model1.predict(Xtr, num_iteration=model1.best_iteration),
            y_test.values,  model1.predict(Xts, num_iteration=model1.best_iteration),
            y_oot.values,   model1.predict(Xoot, num_iteration=model1.best_iteration),
            tw.values, vw.values, ow.values, scheme=self.loss_scheme,
        )
        logs = [{"train_mae": tr_mae1, "test_mae": ts_mae1, "oot_mae": oot_mae1,
                 "loss": loss1, "params": {**p1, "n_estimators": num_br1}}]

        imp = pd.DataFrame({"feature": model1.feature_name(), "importance": model1.feature_importance("gain")})
        self.selected_features = imp[imp["importance"] > 0]["feature"].tolist()
        if not self.selected_features:
            self.selected_features = features

        # ── 第2轮：筛选后特征，Hyperopt 超参搜索 ────────────────────────────
        Xtr2, Xts2, Xoot2 = Xtr[self.selected_features], Xts[self.selected_features], Xoot[self.selected_features]
        train_set2 = lgb.Dataset(Xtr2, y_train, categorical_feature=self.cat_cols, weight=tw, free_raw_data=False)
        test_set2  = lgb.Dataset(Xts2, y_test,  categorical_feature=self.cat_cols, weight=vw, free_raw_data=False)

        logs2 = []

        def objective2(params):
            params["num_leaves"]       = int(params["num_leaves"])
            params["max_depth"]        = int(params["max_depth"])
            num_boost_round2           = int(params.pop("n_estimators"))
            params["lambda_l1"]        = int(params["lambda_l1"])
            params["lambda_l2"]        = int(params["lambda_l2"])
            params["min_data_in_leaf"] = int(params["min_data_in_leaf"])
            m = _train(params, train_set2, num_boost_round2, test_set2)
            loss, tr_mae, ts_mae, oot_mae = _reg_loss_mae(
                y_train.values, m.predict(Xtr2, num_iteration=m.best_iteration),
                y_test.values,  m.predict(Xts2, num_iteration=m.best_iteration),
                y_oot.values,   m.predict(Xoot2, num_iteration=m.best_iteration),
                tw.values, vw.values, ow.values, scheme=self.loss_scheme,
            )
            logs2.append({"train_mae": tr_mae, "test_mae": ts_mae, "oot_mae": oot_mae, "loss": loss,
                          "params": {**params, "n_estimators": num_boost_round2}})
            return {"loss": loss, "status": STATUS_OK}

        space2 = custom_space if custom_space is not None else self.get_default_space(n_data)
        trials2 = _run_hyperopt(objective2, space2, self.n_trials, "LGBM回归 第2轮（超参搜索）")

        idx2    = int(np.argmin([r["loss"] for r in logs2]))
        params2 = trials2.trials[idx2]["misc"]["vals"]
        params2 = {k: v[0] if isinstance(v, list) and len(v) == 1 else v for k, v in params2.items()}
        num_br2 = int(params2.pop("n_estimators", 500))
        params2.update({
            "boosting_type": "gbdt", "objective": "regression", "metric": "mae",
            "num_threads": -1, "verbose": -1, "bagging_freq": 2,
        })
        for k in ["num_leaves", "max_depth", "lambda_l1", "lambda_l2", "min_data_in_leaf"]:
            if k in params2:
                params2[k] = int(params2[k])
        self.model = _train(params2, train_set2, num_br2, test_set2)
        self.best_params = {**params2, "n_estimators": num_br2}
        self.trials_log  = _build_trials_df(logs, logs2)

        if save_dir:
            self.save(save_dir)
        return self

    def predict(self, X: pd.DataFrame) -> np.ndarray:
        feats = self.selected_features or X.columns.tolist()
        return self.model.predict(X[feats], num_iteration=self.model.best_iteration)

    def save(self, save_dir: str) -> None:
        deploy_dir = os.path.join(save_dir, "05_model_deploy")
        os.makedirs(deploy_dir, exist_ok=True)
        payload = {
            "model_type": "lgbm_reg",
            "model": self.model,
            "selected_features": self.selected_features,
            "best_params": self.best_params,
        }
        joblib.dump(payload, os.path.join(deploy_dir, "model.pkl"))
        if not self.trials_log.empty:
            _save_tuning_log(self.trials_log, save_dir)
        print(f"模型已保存至: {deploy_dir}")

    @classmethod
    def load(cls, save_dir: str) -> "LGBMRegressorTrainer":
        payload = joblib.load(os.path.join(save_dir, "05_model_deploy", "model.pkl"))
        trainer = cls()
        trainer.model = payload["model"]
        trainer.selected_features = payload["selected_features"]
        trainer.best_params = payload["best_params"]
        return trainer


class XGBRegressorTrainer:
    """XGBoost 回归训练器"""

    def __init__(self, n_trials: int = 100, use_gpu: bool = False, random_state: int = 2024, loss_scheme: str = "A"):
        self.n_trials = n_trials
        self.use_gpu = use_gpu
        self.random_state = random_state
        self.loss_scheme = loss_scheme  # "A"（默认）或 "B"（train-OOT 差距惩罚）
        self.model = None
        self.best_params = None
        self.selected_features: list = []
        self.trials_log: pd.DataFrame = pd.DataFrame()

    def get_default_params(self) -> dict:
        """返回第一轮固定训练参数。"""
        return {
            "objective": "reg:squarederror", "booster": "gbtree", "eval_metric": "mae",
            "n_estimators": 800, "learning_rate": 0.05, "max_depth": 6,
            "subsample": 0.8, "colsample_bytree": 0.8,
            "min_child_weight": 5, "gamma": 2,
            "reg_alpha": 5, "reg_lambda": 5,
            "early_stopping_rounds": 30, "nthread": -1,
        }

    def get_default_space(self) -> dict:
        """
        返回默认超参数搜索空间字典（第二轮使用）。

        在 Jupyter 中不重启修改参数的用法：
            space = trainer.get_default_space()
            from hyperopt import hp
            space['max_depth'] = hp.choice('max_depth', [3, 4, 5, 6, 7])
            trainer.fit(..., custom_space=space)
        """
        from hyperopt import hp
        return {
            "objective": "reg:squarederror", "booster": "gbtree", "eval_metric": "mae",
            "learning_rate":    hp.quniform("learning_rate", 0.01, 0.3, 0.01),
            "gamma":            hp.quniform("gamma", 0, 10, 1),
            "max_depth":        hp.choice("max_depth", [4, 5, 6, 7, 8]),
            "subsample":        hp.quniform("subsample", 0.7, 1.0, 0.1),
            "colsample_bytree": hp.quniform("colsample_bytree", 0.7, 1.0, 0.1),
            "n_estimators":     800,
            "min_child_weight": hp.quniform("min_child_weight", 1, 30, 1),
            "reg_alpha":        hp.quniform("reg_alpha", 0, 20, 1),
            "reg_lambda":       hp.quniform("reg_lambda", 0, 20, 1),
            "random_state":     self.random_state,
            "tree_method":      "gpu_hist" if self.use_gpu else "hist",
            "early_stopping_rounds": 30,
            "nthread": -1,
        }

    def fit(
        self,
        X_train: pd.DataFrame, y_train: pd.Series,
        X_test:  pd.DataFrame, y_test:  pd.Series,
        X_oot:   pd.DataFrame, y_oot:   pd.Series,
        train_weight: Optional[pd.Series] = None,
        test_weight:  Optional[pd.Series] = None,
        oot_weight:   Optional[pd.Series] = None,
        features: Optional[list] = None,
        custom_space: Optional[dict] = None,
        custom_params: Optional[dict] = None,
        save_dir: Optional[str] = None,
    ) -> "XGBRegressorTrainer":
        try:
            import xgboost as xgb
            from hyperopt import fmin, tpe, hp, Trials, STATUS_OK
        except ImportError as e:
            raise ImportError(f"请安装依赖: pip install xgboost hyperopt — {e}")

        features = features or X_train.columns.tolist()
        Xtr  = X_train[features].reset_index(drop=True)
        Xts  = X_test[features].reset_index(drop=True)
        Xoot = X_oot[features].reset_index(drop=True)
        y_train = y_train.reset_index(drop=True)
        y_test  = y_test.reset_index(drop=True)
        y_oot   = y_oot.reset_index(drop=True)
        tw = train_weight.reset_index(drop=True) if train_weight is not None else pd.Series(np.ones(len(Xtr)))
        vw = test_weight.reset_index(drop=True)  if test_weight  is not None else pd.Series(np.ones(len(Xts)))
        ow = oot_weight.reset_index(drop=True)   if oot_weight   is not None else pd.Series(np.ones(len(Xoot)))

        tree_method = "gpu_hist" if self.use_gpu else "hist"

        # ── 第一轮：固定参数单次训练 ─────────────────────────────────────────────
        print("XGB回归 第1轮（固定参数训练）")
        p1 = self.get_default_params()
        if custom_params:
            p1.update(custom_params)
        p1["tree_method"]   = tree_method
        p1["random_state"]  = self.random_state
        model1 = xgb.XGBRegressor(**p1, verbosity=0)
        model1.fit(Xtr, y_train, sample_weight=tw,
                   eval_set=[(Xts, y_test)], sample_weight_eval_set=[vw], verbose=False)
        loss1, tr_mae1, ts_mae1, oot_mae1 = _reg_loss_mae(
            y_train, model1.predict(Xtr), y_test, model1.predict(Xts), y_oot, model1.predict(Xoot),
            tw.values, vw.values, ow.values, scheme=self.loss_scheme,
        )
        logs = [{"train_mae": tr_mae1, "test_mae": ts_mae1, "oot_mae": oot_mae1, "loss": loss1}]

        # ── 特征筛选 ─────────────────────────────────────────────────────────────
        imp = pd.Series(model1.get_booster().get_score(importance_type="gain")).sort_values(ascending=False)
        self.selected_features = imp.index.tolist() if len(imp) > 0 else features

        Xtr2, Xts2, Xoot2 = Xtr[self.selected_features], Xts[self.selected_features], Xoot[self.selected_features]
        logs2 = []

        def objective2(params):
            reg = xgb.XGBRegressor(**params, verbosity=0)
            reg.fit(Xtr2, y_train, sample_weight=tw,
                    eval_set=[(Xts2, y_test)], sample_weight_eval_set=[vw], verbose=False)
            loss, tr_mae, ts_mae, oot_mae = _reg_loss_mae(
                y_train, reg.predict(Xtr2), y_test, reg.predict(Xts2), y_oot, reg.predict(Xoot2),
                tw.values, vw.values, ow.values, scheme=self.loss_scheme,
            )
            logs2.append({"train_mae": tr_mae, "test_mae": ts_mae, "oot_mae": oot_mae, "loss": loss})
            return {"loss": loss, "status": STATUS_OK}

        space2 = custom_space if custom_space is not None else self.get_default_space()
        trials2 = _run_hyperopt(objective2, space2, self.n_trials, "XGB回归 第2轮")

        best_idx2 = int(np.argmin([r["loss"] for r in logs2]))
        best_t2   = trials2.trials[best_idx2]
        p2 = {k: v[0] if isinstance(v, list) else v for k, v in best_t2["misc"]["vals"].items()}
        p2.update({"objective": "reg:squarederror", "booster": "gbtree", "eval_metric": "mae",
                   "n_estimators": 800, "random_state": self.random_state,
                   "tree_method": tree_method, "early_stopping_rounds": 30, "nthread": -1})
        if "max_depth" in p2:
            from hyperopt.pyll import Apply
            md_space = space2.get("max_depth")
            if hasattr(md_space, "pos_args"):
                candidates = [a.obj for a in md_space.pos_args[1:]]
                p2["max_depth"] = candidates[int(p2["max_depth"])]
            else:
                p2["max_depth"] = int(p2["max_depth"])
        self.model = xgb.XGBRegressor(**p2, verbosity=0)
        self.model.fit(Xtr2, y_train, sample_weight=tw, eval_set=[(Xts2, y_test)],
                       sample_weight_eval_set=[vw], verbose=False)
        self.best_params = p2
        self.trials_log  = _build_trials_df(logs, logs2)

        if save_dir:
            self.save(save_dir)
        return self

    def predict(self, X: pd.DataFrame) -> np.ndarray:
        feats = self.selected_features or X.columns.tolist()
        return self.model.predict(X[feats])

    def save(self, save_dir: str) -> None:
        deploy_dir = os.path.join(save_dir, "05_model_deploy")
        os.makedirs(deploy_dir, exist_ok=True)
        payload = {
            "model_type": "xgboost_reg",
            "model": self.model,
            "selected_features": self.selected_features,
            "best_params": self.best_params,
        }
        joblib.dump(payload, os.path.join(deploy_dir, "model.pkl"))
        if not self.trials_log.empty:
            _save_tuning_log(self.trials_log, save_dir)
        print(f"模型已保存至: {deploy_dir}")

    @classmethod
    def load(cls, save_dir: str) -> "XGBRegressorTrainer":
        payload = joblib.load(os.path.join(save_dir, "05_model_deploy", "model.pkl"))
        trainer = cls()
        trainer.model = payload["model"]
        trainer.selected_features = payload["selected_features"]
        trainer.best_params = payload["best_params"]
        return trainer


class ModelTrainer:
    """统一训练入口，支持 lgbm / xgboost / xgboost_reg / lgbm_reg"""

    def __init__(self, model_type: str = "lgbm", **kwargs):
        self.model_type = model_type
        if model_type == "lgbm":
            self._trainer = LGBMTrainer(**kwargs)
        elif model_type == "xgboost":
            self._trainer = XGBTrainer(**kwargs)
        elif model_type == "xgboost_reg":
            self._trainer = XGBRegressorTrainer(**kwargs)
        elif model_type == "lgbm_reg":
            self._trainer = LGBMRegressorTrainer(**kwargs)
        else:
            raise ValueError(f"model_type 须为 'lgbm' / 'xgboost' / 'xgboost_reg' / 'lgbm_reg'，不支持 '{model_type}'")

    def get_default_params(self, **kwargs) -> dict:
        """获取当前模型类型的第一轮固定训练参数，可在 notebook 中查看后传给 fit(custom_params=...)"""
        return self._trainer.get_default_params(**kwargs)

    def get_default_space(self, **kwargs) -> dict:
        """获取当前模型类型的默认参数空间，可在 notebook 中修改后传给 fit(custom_space=...)"""
        return self._trainer.get_default_space(**kwargs)

    def fit(self, X_train, y_train, X_test, y_test, X_oot, y_oot,
            custom_space=None, save_dir=None, **kwargs) -> "ModelTrainer":
        self._trainer.fit(X_train, y_train, X_test, y_test, X_oot, y_oot,
                          custom_space=custom_space, save_dir=save_dir, **kwargs)
        return self

    def predict_proba(self, X: pd.DataFrame) -> np.ndarray:
        if self.model_type in ("xgboost_reg", "lgbm_reg"):
            raise TypeError(f"{self.model_type} 是回归模型，请使用 predict()")
        return self._trainer.predict_proba(X)

    def predict(self, X: pd.DataFrame) -> np.ndarray:
        if self.model_type not in ("xgboost_reg", "lgbm_reg"):
            raise TypeError(f"{self.model_type} 是分类模型，请使用 predict_proba()")
        return self._trainer.predict(X)

    def save(self, save_dir: str) -> None:
        """保存模型文件、入模特征和参数元信息到 save_dir"""
        self._trainer.save(save_dir)

    @classmethod
    def load(cls, save_dir: str) -> "ModelTrainer":
        """从 save_dir/model.pkl 恢复模型，自动识别模型类型"""
        payload = joblib.load(os.path.join(save_dir, "05_model_deploy", "model.pkl"))
        model_type = payload["model_type"]
        instance = cls.__new__(cls)
        instance.model_type = model_type
        if model_type == "lgbm":
            instance._trainer = LGBMTrainer.load(save_dir)
        elif model_type == "xgboost":
            instance._trainer = XGBTrainer.load(save_dir)
        elif model_type == "xgboost_reg":
            instance._trainer = XGBRegressorTrainer.load(save_dir)
        elif model_type == "lgbm_reg":
            instance._trainer = LGBMRegressorTrainer.load(save_dir)
        else:
            raise ValueError(f"未知 model_type: {model_type}")
        return instance

    @property
    def selected_features(self):
        return self._trainer.selected_features

    @property
    def best_params(self):
        return self._trainer.best_params

    @property
    def trials_log(self):
        return self._trainer.trials_log
